from tempfile import NamedTemporaryFile

from django.db.models.functions import Concat, Trim
from django.db.models import Value, Subquery, OuterRef
from django.http import HttpResponse

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import JSONParser
from rest_framework.permissions import IsAuthenticated, IsAdminUser

<<<<<<< HEAD
from drf_spectacular.utils import extend_schema
=======
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiExample, OpenApiResponse
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import NamedStyle, Side, Font, Border, Alignment

from integrator.apps.parsers import NestedMultipartParser

from spares.api.serializers import *

from spares.models import SpareModel, PartNumberModel, SparePNModel

<<<<<<< HEAD
#Запчасти
=======
@extend_schema(
    tags = ['Запчасти (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class SparesListAPIView(APIView):
    permission_classes = [IsAuthenticated,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список запчастей (Done)'],
=======
        summary = 'Список запчастей',
        description = '<ol><li>"id" - Идентификатор запчасти</li><li>"spare" - Наименование и серийный номер запчасти</li>\
        <li>"name" - Наименование запчасти</li><li>"sn" - Серийный номер запчасти</li>\
        <li>"description" - Описание запчасти</li><li>"barcode" - Ссылка на шрих-код запчасти</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = SparesSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):
        spares = SpareModel.objects.all()\
            .annotate(
                spare = Trim(Concat('name', Value(' (S/n: '), 'sn', Value(')')))
            )

        serializer = SparesSerializer(spares, many = True)
        return Response(serializer.data, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Запчасти (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class SparesListAdminAPIView(APIView):
    permission_classes = [IsAdminUser,]
    parser_classes = [JSONParser, NestedMultipartParser]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список запчастей. Администрирование (Done)'],
=======
        summary = 'Список запчастей (Администрирование)',
        description = '<ol><li>"id" - Идентификатор запчасти</li>\
        <li>"name" - Наименование запчасти</li><li>"sn" - Серийный номер запчасти</li>\
        <li>"description" - Описание запчасти</li><li>"pn" - Один из партномеров запчасти</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = SparesAdminSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):
        spares = SpareModel.objects.all()\
            .annotate(
                pn = Subquery(SparePNModel.objects.filter(spare_id = OuterRef('id'))[:1].values('number__number'))
            )

        serializer = SparesAdminSerializer(spares, many = True)
        return Response(serializer.data, status = status.HTTP_200_OK)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список запчастей. Администрирование. Создание (Done)'],
        request = EditSpareSerializer()
=======
        summary = 'Добавление запчасти',
        description = '<ol><li>"name" - Наименование запчасти</li><li>"sn" - Серийный номер запчасти</li>\
        <li>"description" - Описание запчасти</li><li>"pnspare" - Список партномеров, где<ul>\
        <li>"id" - Идентификатор привязки партномера к запчасти <b>(Не используется в данном методе, указывать не нужно)</b></li>\
        <li>"number" - Идентификатор партномера</li><li>"DELETE" - True, если необходимо убрать привязку партномера к запчасти</li></ul></li></ol>',
        request = EditSpareSerializer(),
        responses = {(201, 'application/json'): OpenApiResponse(response = EditSpareSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EditSpareSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Запчасти (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class SparesExcelAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список запчастей. Администрирование. Экспорт (Done)'],
=======
        summary = 'Экспорт списка запчастей в Excel',
        description = 'Response - файл с данными',
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):
        spares = SparesListAdminAPIView().get(request = request).data

        stream = ExportItems(spares, ['Наименование', 'Серийный номер', 'Описание', 'Партномер'], 'ЗИП')

        response = HttpResponse(content = stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=zip.xlsx'
        return response

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список запчастей. Администрирование. Импорт (Done)'],
        request = SparesImportSerializer()
=======
        summary = 'Импорт списка запчастей из Excel',
        description = '"excel" - xlsx-файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b>',
        request = SparesImportSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = LoadedSparesLoadedSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        wb = load_workbook(filename = request.data.get('excel').file)

        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        start = 0
        end = ws.max_row
        records = []

        for m in range(3 , end + 1):
            sp, created = SpareModel.objects.update_or_create(sn = ws.cell(m, 3).value, defaults = {'name': ws.cell(m, 2).value, 'description': ws.cell(m, 4).value})

            if ws.cell(m, 5).value:
                temp = ws.cell(m, 5).value.splitlines()
            if temp:
                for ln in temp:
                    nm, created = PartNumberModel.objects.get_or_create(number = ln)
                    records.append({'spare': sp, 'number': nm})

        list = [SparePNModel(**vals) for vals in records]
        SparePNModel.objects.bulk_create(list, ignore_conflicts = True)
        return Response(records, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Запчасти (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class SpareEditAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список запчастей. Администрирование. Изменение (Done)'],
        request = EditSpareSerializer()
=======
        summary = 'Изменение запчасти',
        description = '<ol><li>"name" - Наименование запчасти</li><li>"sn" - Серийный номер запчасти</li>\
        <li>"description" - Описание запчасти</li><li>"pnspare" - Список партномеров, где<ul>\
        <li>"id" - Идентификатор привязки партномера к запчасти <b>(Указывется в случае изменения / удаления уже существующей привязки)</b></li>\
        <li>"number" - Идентификатор партномера</li><li>"DELETE" - True, если необходимо убрать привязку партномера к запчасти</li></ul></li></ol>',
        parameters = [
            OpenApiParameter(name = 'spare_id', description = 'Идентификатор запчасти', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = EditSpareSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EditSpareSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def put(self, request, spare_id, *args, **kwargs):
        spare = GetSpare(spare_id)
        serializer = EditSpareSerializer(spare, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список запчастей. Администрирование. Удаление (Done)'],
=======
        summary = 'Удаление запчасти',
        parameters = [
            OpenApiParameter(name = 'spare_id', description = 'Идентификатор запчасти', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элемент удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Элемент удален'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def delete(self, request, spare_id, *args, **kwargs):
        spare = GetSpare(spare_id)
        spare.delete()
        return Response({'message': 'Элемент удален'}, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Запчасти (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class SparesDeleteAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список запчастей. Администрирование. Удаление нескольких (Done)'],
        request = SparesDeleteSerializer(),
=======
        summary = 'Удаление списка запчастей',
        description = 'Используется метод POST, т.к. методом DELETE невозможно отправить тело запроса\
        <p>"id" - Список идентификаторов запчастей { "id": [1, 2, 3] }</p>',
        request = SparesDeleteSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элементы удалены'}, examples = [OpenApiExample('Пример', value = {'message': 'Элементы удалены'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = SparesDeleteSerializer({'id': request.data.getlist('id')})
        if serializer:
            SpareModel.objects.filter(id__in = serializer.data.get('id')).delete()
            return Response({'message': 'Элементы удалены'}, status = status.HTTP_200_OK)
        return Response({'error': 'Некорректные данные'}, status=status.HTTP_400_BAD_REQUEST)

def GetSpare(spare_id):
    try:
        return SpareModel.objects.get(id = spare_id)
    except:
        return Response({'message': 'Объект не найден'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
#Партномера
=======
@extend_schema(
    tags = ['Партномера (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class PartNumberListAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список партномеров. Администрирование (Done)'],
    )
    def get(self, request, *args, **kwargs):
        pns = PartNumberModel.objects.all()

=======
        summary = 'Список партномеров',
        description = '<ol><li>"id" - Идентификатор партномера</li><li>"number" - Партномер</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = PartNumberSerializer(many = True))}
    )
    def get(self, request, *args, **kwargs):
        pns = PartNumberModel.objects.all()
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
        serializer = PartNumberSerializer(pns, many = True)
        return Response(serializer.data, status = status.HTTP_200_OK)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список партномеров. Администрирование. Создание (Done)'],
        request = PartNumberSerializer()
=======
        summary = 'Добавление партномера',
        description = '"number" - Партномер',
        request =  PartNumberSerializer(),
        responses = {(201, 'application/json'): OpenApiResponse(response = PartNumberSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = PartNumberSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Партномера (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class PartNumberEditAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список партномеров. Администрирование. Изменение (Done)'],
        request = PartNumberSerializer()
=======
        summary = 'Изменение партномера',
        description = '"number" - Партномер',
        parameters = [
            OpenApiParameter(name = 'pn_id', description = 'Идентификатор партномера', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = PartNumberSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = PartNumberSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def put(self, request, pn_id, *args, **kwargs):
        pn = GetPartNumber(pn_id)
        serializer = PartNumberSerializer(pn, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список партномеров. Администрирование. Удаление (Done)'],
=======
        summary = 'Удаление партномера',
        parameters = [
            OpenApiParameter(name = 'pn_id', description = 'Идентификатор партномера', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элемент удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Элемент удален'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def delete(self, request, pn_id, *args, **kwargs):
        pn = GetPartNumber(pn_id)
        pn.delete()
        return Response({'message': 'Элемент удален'}, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Партномера (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class PartNumbersDeleteAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список партномеров. Администрирование. Удаление нескольких (Done)'],
        request = PartNumbersDeleteSerializer(),
=======
        summary = 'Удаление списка партномеров',
        description = 'Используется метод POST, т.к. методом DELETE невозможно отправить тело запроса\
        <p>"id" - Список идентификаторов партномеров { "id": [1, 2, 3] }</p>',
        request =  PartNumbersDeleteSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элементы удалены'}, examples = [OpenApiExample('Пример', value = {'message': 'Элементы удалены'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = PartNumbersDeleteSerializer({'id': request.data.getlist('id')})
        if serializer:
            PartNumberModel.objects.filter(id__in = serializer.data.get('id')).delete()
            return Response({'message': 'Элементы удалены'}, status = status.HTTP_200_OK)
        return Response({'error': 'Некорректные данные'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Партномера (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class PartNumbersExcelAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список партномеров. Администрирование. Экспорт (Done)'],
=======
        summary = 'Экспорт списка партномеров в Excel',
        description = 'Response - файл с данными',
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):
        pns = PartNumberListAPIView().get(request = request).data
        fields = ['Партномер',]

        stream = ExportItems(pns, fields, 'Партномера')

        response = HttpResponse(content = stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=partnumbers.xlsx'
        return response

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список партномеров. Администрирование. Импорт (In progress)'],
        request = SparesImportSerializer()
=======
        summary = 'Импорт списка партномеров из Excel',
        description = '"excel" - xlsx-файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b>',
        request = SparesImportSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = PartNumberSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        wb = load_workbook(filename = request.data.get('excel').file)

        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        start = 0
        end = ws.max_row
        records = []

        for m in range(3 , end + 1):
            records.append({'number': ws.cell(m, 2).value})

        list = [PartNumberModel(**vals) for vals in records]
        PartNumberModel.objects.bulk_create(list, ignore_conflicts = True)
        return Response(records, status = status.HTTP_200_OK)

def GetPartNumber(pn_id):
    try:
        return PartNumberModel.objects.get(id = pn_id)
    except:
        return Response({'message': 'Объект не найден'}, status = status.HTTP_400_BAD_REQUEST)

def ExportItems(items, fields, title):
    wb = Workbook()
    with NamedTemporaryFile(delete = True) as tmp:
        ws = wb.active
        ws.title = title

        bd = Side(style = 'thin', color = '000000')

        general = NamedStyle(name = 'general')
        general.font = Font(name = 'Times New Roman', size = 14)
        general.border = Border(left = bd, top = bd, right = bd, bottom = bd)
        general.alignment = Alignment(vertical = 'center', wrap_text = True)
        wb.add_named_style(general)

        header = NamedStyle(name = 'header')
        header.font = Font(name = 'Times New Roman', size = 14, bold = True)
        header.border = Border(left = bd, top = bd, right = bd, bottom = bd)
        header.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
        wb.add_named_style(header)

        ws.column_dimensions['A'].width = 10

        ws['A2'] = '#'
        ws['A2'].style = 'header'
        col = 2

        for field in fields:
            cell = ws.cell(row = 2, column = col, value = field)
            ws.column_dimensions[get_column_letter(col)].width = 20
            cell.style = 'header'
            col = col + 1

        for count, item in enumerate(items):
            cell = ws.cell(row = count + 3, column = 1, value = count + 1)
            cell.style = 'general'
            col = 2
            for field in item.keys():
                if not field == 'id':
                    cell = ws.cell(row = count + 3, column = col, value = str(item[field] or ''))
                    cell.style = 'general'
                    col = col + 1

        wb.save(tmp.name)
        tmp.seek(0)
        return tmp.read()
