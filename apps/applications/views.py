import os
from datetime import datetime

from django.shortcuts import render, redirect, HttpResponse
from django.views.generic.edit import FormView
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_POST
from django.conf import settings

from django.db.models import Subquery, OuterRef, Value, Q, F, Case, When, IntegerField, CharField
from django.db.models.functions import Concat, Trim

from applications.models import AppPriorityModel, StatusModel, ApplicationModel, EquipmentConfigModel, AppHistoryModel

from equipments.models import EquipmentModel

from contracts.models import ContractEquipmentModel

from accounts.models import OrganizationContactModel

from spares.models import SpareModel

from applications.api.serializers import AppPrioritySerializer, AppStatusSerializer

from applications.api.api import *
from integrator.apps.api_services import *
from accounts.api.api import ContactsListAPIView

from applications.forms import ApplicationForm, AppDocumentsFormset, StaffApplicationForm, ClientApplicationForm
from applications.forms import AppEditCommentForm, AppDeleteCommentForm

from applications.permissions import IsAppAdmin, IsAppStaff
from integrator.apps.functions import is_admin_or_engineer, get_prms_from_ids

@login_required
@permission_required('applications.add_applicationmodel')
def AddApplicationView(request): #Создание заявки
    User = get_user_model()

    if not request.user.groups.filter(name = 'Администратор').exists() and not request.user.groups.filter(name = 'Инженер').exists():
        form = ApplicationForm(request.user)
    else:
        form = ApplicationForm()

    doc_formset = AppDocumentsFormset()

    if request.method == "POST":
        from django.http import JsonResponse, HttpResponse
        from contracts.models import ContractModel

        if 'problem' not in request.POST:
            #Q(contract__client_id = request.POST.get('client')
            client = User.objects.get(id = request.POST.get('client'))
            equipments = ContractEquipmentModel.objects.filter(Q(contract__client__organization_id = client.organization.id) & Q(contract__enddate__gt = datetime.now().date()) & Q(Q(sn__icontains = request.POST.get('sn')) | Q(equipment__model__name__icontains = request.POST.get('sn')))).annotate(conf = Value('')).values('equipment__id', 'equipment__brand__name', 'equipment__model__name', 'sn')
            if request.user.groups.filter(name = 'Администратор').exists() or request.user.groups.filter(name = 'Инженер').exists():
                if len(equipments) == 1:
                    confs = EquipmentConfigModel.objects.filter(equipment__id = equipments[0]['equipment__id'])
                    for doc in confs:
                        doc.filename = os.path.basename(doc.document.name)
                        equipments[0]['conf'] += '<p><a target="_blank" href="' + settings.MEDIA_URL + doc.document.name + '">' + doc.filename + '</a></p>'

            return JsonResponse(list(equipments), safe = False)

        if not request.user.groups.filter(name = 'Администратор').exists() and not request.user.groups.filter(name = 'Инженер').exists():
            form = ApplicationForm(request.user, request.POST)
        else:
            form = ApplicationForm(None, request.POST)

        from applications.models import AppStatusModel

        if form.is_valid():
            app = form.save(commit = False)

            if not request.user.groups.filter(name = 'Администратор').exists() and not request.user.groups.filter(name = 'Инженер').exists():
                app.client = request.user
            if request.user.groups.filter(name = 'Инженер').exists():
                app.engineer = request.user
                app.status = StatusModel.objects.get(id = 6)
            else:
                app.status = StatusModel.objects.get(id = 1)
            app.creator = request.user
            eq_list = request.POST.get('equipment')[0:-1].split(' (S/n: ')
            app.equipment = ContractEquipmentModel.objects.get(sn = eq_list[1])
            doc_formset = AppDocumentsFormset(request.POST, request.FILES, instance = app)

            if doc_formset.is_valid():
                app.save()
                doc_formset.save()
                AppStatusModel.objects.create(status = app.status, application = app)
                AppHistoryModel.objects.create(type = 1, text = 'Заявка создана, присвоен статус "' + str(app.status) + '"', application = app, author = request.user)

                from django.template.loader import render_to_string
                from django.core.mail import EmailMessage
                import telegram

                text = render_to_string('applications/mail.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'type': 'add'})

                mail = EmailMessage('Создание заявки № ' + str(app.id), text, settings.EMAIL_HOST_USER, [app.contact.email])
                mail.content_subtype = "html"
                try:
                    mail.send()
                except Exception:
                    AppHistoryModel.objects.create(type = 3, text = 'Не удалось отправить сообщение о создании заявки на адрес электронной почты ' + app.contact.email, application = app, author = request.user)
                else:
                    AppHistoryModel.objects.create(type = 3, text = 'Отправлено сообщение о создании заявки на адрес электронной почты ' + app.contact.email, application = app, author = request.user)

                if app.status.id == 6: # Сообщение инженеру
                    text = render_to_string('applications/mail.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'status': app.engineer, 'type': 'engineer'})
                    mail = EmailMessage('Назначен инженер', text, settings.EMAIL_HOST_USER, [app.engineer.email])
                    mail.content_subtype = "html"
                    try:
                        mail.send()
                    except Exception:
                        AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение о назначении инженера "' + str(app.engineer) + '" на адрес электронной почты ' + app.engineer.email, application = app, author = request.user)
                    else:
                        AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение о назначении инженера "' + str(app.engineer) + '" на адрес электронной почты ' + app.engineer.email, application = app, author = request.user)

                text = render_to_string('applications/telegram.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'type': 'add'})
                telegram_settings = settings.TELEGRAM
                bot = telegram.Bot(token = telegram_settings['bot_token'])
                try:
                    bot.send_message(chat_id = telegram_settings['channel_id'], text = text, parse_mode = telegram.ParseMode.HTML)
                except Exception:
                    AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение о создании заявки в телеграм-канал', application = app, author = request.user)
                else:
                    AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение о создании заявки в телеграм-канал', application = app, author = request.user)

                return JsonResponse({'message': 1})
            else:
                error_text = {}
                for doc in doc_formset:
                    if not doc.is_valid():
                        error_text[request.FILES[doc.prefix + '-document'].name] = doc.errors['document']

                return JsonResponse(error_text)

    if not request.user.groups.filter(name = 'Администратор').exists() and not request.user.groups.filter(name = 'Инженер').exists():
        context = {'form': form, 'doc_formset': doc_formset, 'show': False}
    else:
        context = {'form': form, 'doc_formset': doc_formset, 'show': True, 'admin': request.user.groups.filter(name = 'Администратор').exists()}

    return render(request, 'applications/add.html', context)

@login_required
def EditApplicationView(request, application_id): #Редактирование заявки
    if not request.user.groups.filter(name = 'Администратор').exists() and not request.user.groups.filter(name = 'Инженер').exists():
        return redirect('login')

    from applications.models import AppStatusModel
    from applications.forms import EditApplicationForm, AppStatusFormset, AppSpareFormset, EditAppDocumentsFormset


    data = EditApplicationAPIView.as_view()(request = request, application_id = application_id).data
    app = GetApplication(application_id)
    confs = EquipmentConfigModel.objects.filter(equipment = app.equipment.equipment).annotate(filename = Value(''))
    for doc in confs:
        doc.filename = os.path.basename(doc.document.name)

    form = EditApplicationForm(instance = app)
    formset = AppStatusFormset(instance = app)
    document = EditAppDocumentsFormset(instance = app)
    comment = AppEditCommentForm()
    spares = AppSpareFormset(instance = app)

    if request.method == "POST":
        import telegram
        from django.template.loader import render_to_string
        from django.core.mail import EmailMessage

        telegram_settings = settings.TELEGRAM
        bot = telegram.Bot(token = telegram_settings['bot_token'])

        if spares.prefix + '-TOTAL_FORMS' in request.POST:
            spares = AppSpareFormset(request.POST, instance = app)
            if spares.is_valid():
                for spare in spares:
                    appspare = spare.save(commit = False)
                    parent = SpareModel.objects.get(id = spare.cleaned_data.get('spare').id)

                    if not spare.instance.pk:
                        appspare.author = request.user

                    parent.save()
                spares.save()
                return redirect('edit-application', application_id = application_id)
        else:

            form = EditApplicationForm(request.POST, instance = app)
            formset = AppStatusFormset(request.POST, instance = app)
            document = EditAppDocumentsFormset(request.POST, request.FILES, instance = app)
            if not request.POST.get('appstatuses-0-status'):
                updated_request = request.POST.copy()
                updated_request.update({'appstatuses-0-status': app.status})
                formset = AppStatusFormset(updated_request, instance = app)

            if document.is_valid():
                document.save()

            if form.is_valid():
                # if not request.user.groups.filter(name = 'Администратор').exists() and 'engineer' in form.changed_data:
                #     return redirect('login')

                app = form.save(commit = False)
                from applications.models import AppStatusModel
                st = AppStatusModel.objects.filter(application = app).values('status_id', 'status__name')[0]

                if not app.changed: #Замена оборудования
                    eq_list = request.POST.get('equipment')[0:-1].split(' (S/n: ')
                    new_equip = ContractEquipmentModel.objects.get(sn = eq_list[1])
                    if new_equip != app.equipment:
                        app.equipment = new_equip
                        app.changed = True

                if formset.is_valid():
                    status = formset.save(commit = False)
                    if status and app.status_id != status[0].status_id or app.status_id == 1 and 'engineer' in form.changed_data and app.engineer:
                        if app.status_id == 1 and 'engineer' in form.changed_data and app.engineer:
                            app.status = StatusModel.objects.get(id = 6)
                            AppStatusModel.objects.create(application_id = app.id, status_id = 6)
                        else:
                            app.status = status[0].status
                            status[0].pk = None
                            AppStatusModel.objects.create(application_id = app.id, status = status[0].status)

                        AppHistoryModel.objects.create(type = 1, text = 'Статус заявки изменен на "' + str(app.status) + '"', application = app, author = request.user)

                        text = render_to_string('applications/mail.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'status': app.status, 'type': 'edit'})
                        mail = EmailMessage('Изменение статуса заявки', text, settings.EMAIL_HOST_USER, [app.contact.email])
                        mail.content_subtype = "html"
                        try:
                            mail.send()
                        except Exception:
                            AppHistoryModel.objects.create(type = 3, text = 'Не удалось отправить сообщение об изменении статуса заявки на "' + str(app.status) + '" на адрес электронной почты ' + app.contact.email, application = app, author = request.user)
                        else:
                            AppHistoryModel.objects.create(type = 3, text = 'Отправлено сообщение об изменении статуса заявки на "' + str(app.status) + '" на адрес электронной почты ' + app.contact.email, application = app, author = request.user)

                        if app.engineer: #Сообщение инженеру
                            text = render_to_string('applications/mail.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'status': app.status, 'type': 'status'})
                            mail = EmailMessage('Изменение статуса заявки', text, settings.EMAIL_HOST_USER, [app.engineer.email])
                            mail.content_subtype = "html"
                            try:
                                mail.send()
                            except Exception:
                                AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение об изменении статуса заявки "' + str(app.status) + '" на адрес электронной почты ' + app.engineer.email, application = app, author = request.user)
                            else:
                                AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение об изменении статуса заявки "' + str(app.status) + '" на адрес электронной почты ' + app.engineer.email, application = app, author = request.user)

                        text = render_to_string('applications/telegram.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'status': app.status, 'type': 'edit'})
                        try:
                            bot.send_message(chat_id = telegram_settings['channel_id'], text = text, parse_mode = telegram.ParseMode.HTML)
                        except Exception:
                            AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение об изменении статуса заявки на "' + str(app.status) + '" в телеграм-канал', application = app, author = request.user)
                        else:
                            AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение об изменении статуса заявки на "' + str(app.status) + '" в телеграм-канал', application = app, author = request.user)

                if 'priority' in form.changed_data:
                    AppHistoryModel.objects.create(type = 2, text = 'Изменен приоритет заявки на "' + str(app.priority) + '"', application = app, author = request.user)

                    if app.engineer: #Сообщение инженеру
                        text = render_to_string('applications/mail.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'status': app.priority, 'type': 'priority'})
                        mail = EmailMessage('Изменение приоритета заявки', text, settings.EMAIL_HOST_USER, [app.engineer.email])
                        mail.content_subtype = "html"
                        try:
                            mail.send()
                        except Exception:
                            AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение об изменении приоритета заявки "' + str(app.priority) + '" на адрес электронной почты ' + app.engineer.email, application = app, author = request.user)
                        else:
                            AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение об изменении приоритета заявки "' + str(app.priority) + '" на адрес электронной почты ' + app.engineer.email, application = app, author = request.user)

                    text = render_to_string('applications/telegram.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'status': app.priority, 'type': 'priority'})
                    try:
                        bot.send_message(chat_id = telegram_settings['channel_id'], text = text, parse_mode = telegram.ParseMode.HTML)
                    except Exception:
                        AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение об изменении приоритета заявки на "' + str(app.priority) + '" в телеграм-канал', application = app, author = request.user)
                    else:
                        AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение об изменении приоритета заявки на "' + str(app.priority) + '" в телеграм-канал', application = app, author = request.user)

                if 'engineer' in form.changed_data and app.engineer:
                    AppHistoryModel.objects.create(type = 2, text = 'Назначен инженер "' + str(app.engineer) + '"', application = app, author = request.user)

                    text = render_to_string('applications/mail.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'status': app.engineer, 'type': 'engineer'})
                    mail = EmailMessage('Назначен инженер', text, settings.EMAIL_HOST_USER, [app.engineer.email])
                    mail.content_subtype = "html"
                    try:
                        mail.send()
                    except Exception:
                        AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение о назначении инженера "' + str(app.engineer) + '" на адрес электронной почты ' + app.engineer.email, application = app, author = request.user)
                    else:
                        AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение о назначении инженера "' + str(app.engineer) + '" на адрес электронной почты ' + app.engineer.email, application = app, author = request.user)

                    text = render_to_string('applications/telegram.html', {'id': app.id, 'url': request.build_absolute_uri(app.get_absolute_url()), 'status': app.engineer, 'type': 'engineer'})
                    try:
                        bot.send_message(chat_id = telegram_settings['channel_id'], text = text, parse_mode = telegram.ParseMode.HTML)
                    except Exception:
                        AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение о назначении инженера "' + str(app.engineer) + '" в телеграм-канал', application = app, author = request.user)
                    else:
                        AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение о назначении инженера "' + str(app.engineer) + '" в телеграм-канал', application = app, author = request.user)

                app.save()

            return redirect('edit-application', application_id = application_id)

    context = {'data': data, 'form': form, 'formset': formset, 'comment': comment, 'spares': spares, 'document': document, 'confs': confs, 'admin': request.user.groups.filter(name = 'Администратор').exists()}
    return render(request, 'applications/edit.html', context)

@login_required
@require_POST
def EqModelsView(request): #Выбрать модели оборудования по бренду
    if request.method == 'POST':
        from django.http import JsonResponse
        from applications.models import ModelModel

        pk = request.POST.get('pk')

        eq_models = ModelModel.objects.filter(brand_id = pk).values()

        return JsonResponse(list(eq_models), safe = False)

@login_required
@require_POST
def GetEquipmentsInEdit(request):#Список оборудования в редактировании заявки
    if request.method == 'POST':
        from django.http import JsonResponse
        #Q(contract__client_id = request.POST.get('client')

        app = ApplicationModel.objects.get(pk = request.POST.get('pk'))

        equipments = ContractEquipmentModel.objects.filter(Q(contract_id = app.equipment.contract.id) & Q(Q(sn__icontains = request.POST.get('sn')) | Q(equipment__model__name__icontains = request.POST.get('sn')))).annotate(conf = Value('')).values('equipment__id', 'equipment__brand__name', 'equipment__model__name', 'sn')
        if request.user.groups.filter(name = 'Администратор').exists() or request.user.groups.filter(name = 'Инженер').exists():
            if len(equipments) == 1:
                confs = EquipmentConfigModel.objects.filter(equipment__id = equipments[0]['equipment__id'])
                for doc in confs:
                    doc.filename = os.path.basename(doc.document.name)
                    equipments[0]['conf'] += '<p class="pl-30"><a target="_blank" href="' + settings.MEDIA_URL + doc.document.name + '">' + doc.filename + '</a></p>'

        return JsonResponse(list(equipments), safe = False)

@login_required
def DetailsApplicationView(request, application_id): #Детали о заявке
    from applications.forms import EditAppDocumentsFormset, AppCommentForm

    data = ApplicationDetailsAPIView().get(request = request, application_id = application_id).data
    app = GetApplication(application_id)

    if not data['application']['organization_id'] == request.user.organization_id and not data['permissions']['is_staff']:
        return redirect('login')

    document = EditAppDocumentsFormset(instance = app)
    comment = AppCommentForm()

    if request.method == 'POST':
        document = EditAppDocumentsFormset(request.POST, request.FILES, instance = app)
        if document.is_valid():
            document.save()
            AppHistoryModel.objects.create(type = 7, text = 'Добавлен(-ы) файл(-ы)', application_id = application_id, author = request.user)
            return redirect('app-details', application_id = application_id)

    context = {'app': app, 'data': data, 'comment': comment, 'document': document, 'admin': data['permissions']['is_admin']}

    return render(request, 'applications/details.html', context)

@login_required
def ApplicationHistoryView(request, application_id, type): #Подробная история действий по заявке
    if not is_admin_or_engineer(request.user):
        return redirect('login')

    if type:
        history = ApplicationLogsAPIView().get(request = request, application_id = application_id).data
    else:
        history = ApplicationHistoryAPIView().get(request = request, application_id = application_id).data

    context = {'application_id': application_id, 'history': history}
    return render(request, 'applications/history.html', context)

@login_required
def EditCommentView(request, comment_id, link_type): #Редактирование комментария
    comment = GetComment(comment_id)
    if not is_admin_or_engineer(request.user) and not request.user == comment.author:
        return redirect('login')

    form = AppEditCommentForm(instance = comment)
    link = reverse('app-history', kwargs = {'application_id': comment.application_id, 'type': link_type})
    context = {'form': form, 'media': True, 'method': 'PUT', 'action': reverse('edit-app-comm-api', kwargs = {'comment_id': comment_id}), 'link': link}
    return render(request, 'forms/edit.html', context)

@login_required
def DeleteCommentView(request, comment_id, link_type): #Удаление комментария
    comment = GetComment(comment_id)
    if not is_admin_or_engineer(request.user) and not request.user == comment.author:
        return redirect('login')

    form = AppDeleteCommentForm(instance = comment)
    link = reverse('app-history', kwargs = {'application_id': comment.application_id, 'type': link_type})
    context = {'form': form, 'action': reverse('edit-app-comm-api', kwargs = {'comment_id': comment_id}), 'link': link}
    return render(request, 'forms/delete.html', context)

def SnEqView(request, sn): #Выбрать оборудование по серийному номеру
    from django.http import JsonResponse

    if request.user.groups.filter(name = 'Администратор').exists() or request.user.groups.filter(name = 'Инженер').exists():
        equipments = EquipmentModel.objects.filter(sn__icontains = sn).values('id', 'brand__name', 'model__name', 'sn')
    else:
        equipments = EquipmentModel.objects.filter(sn__icontains = sn, client_id = request.user.id).values('id', 'brand__name', 'model__name', 'sn')

    return JsonResponse(list(equipments), safe = False)

@login_required
@require_POST
def GetContactsView(request): #Выбрать контакты для заявки
    if request.method == 'POST':
        from django.http import JsonResponse, HttpResponse

        User = get_user_model()

        client = request.POST.get('client')
        if not request.user.groups.filter(name = 'Администратор').exists() and not request.user.groups.filter(name = 'Инженер').exists():
            client = request.user.id
        organization = User.objects.get(id = client)

        contacts = ContactsListAPIView().get(request, organization.organization_id).data
        return JsonResponse(list(contacts), safe = False)

@login_required
@require_POST
def GetOrganizationView(request): #Выбрать организацию по пользователю
    if request.method == 'POST':
        from django.http import JsonResponse

        User = get_user_model()
        client = User.objects.get(id = request.POST.get('pk'))

        return JsonResponse({'organization': client.organization})

@login_required
@require_POST
def GetSpareModels(request): #Выбор моделей ЗИП по бренду в админке
    if request.method == 'POST':
        from django.http import JsonResponse
        from applications.models import SpareNameModel

        brand = request.POST.get('brand')
        models = SpareNameModel.objects.filter(brand__pk = brand).values()

        return JsonResponse(list(models), safe = False)

@login_required
@require_POST
def GetModelPNs(request): #Выбор партномеров по модели ЗИП в админке
    if request.method == 'POST':
        from django.http import JsonResponse
        from applications.models import SparePartNumberModel

        model = request.POST.get('model')
        pns = SparePartNumberModel.objects.filter(model__pk = model).values_list('number', flat = True)

        return JsonResponse(list(pns), safe = False)

@login_required
@require_POST
def GetSparesList(request): #Выбор ЗИПов в редактировании заявки
    if request.method == 'POST':
        from django.http import JsonResponse
        from spares.models import SparePNModel

        spare = request.POST.get('spare')
        spare_pns = SparePNModel.objects.filter(number__number__icontains = spare).values_list('spare__pk', flat = True)
        spares = SpareModel.objects.filter(Q(name__icontains = spare) | Q(sn__icontains = spare) | Q(pk__in = spare_pns)).values('name', 'sn', 'pk')

        return JsonResponse(list(spares), safe = False)

@login_required
def EquipmentExcelView(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    from openpyxl.styles import NamedStyle, Side, Font, Border, Alignment
    from contracts.models import ContractModel

    contracts = ContractModel.objects.filter(client__organization_id = request.user.organization_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Оборудование'

    general = NamedStyle(name = 'general')
    bd = Side(style = 'thin', color = '000000')
    general.font = Font(name = 'Times New Roman', size = 14)
    general.border = Border(left = bd, top = bd, right = bd, bottom = bd)
    general.alignment = Alignment(vertical = 'center', wrap_text = True)
    wb.add_named_style(general)

    header = NamedStyle(name = 'header')
    header.font = Font(name = 'Times New Roman', size = 14, bold = True)
    header.border = Border(left = bd, top = bd, right = bd, bottom = bd)
    header.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
    wb.add_named_style(header)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60

    ws['A1'] = 'Список оборудования'
    ws['A1'].font = Font(name = 'Times New Roman', size = 16,  bold = True)
    ws['A1'].alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
    ws.merge_cells('A1:B1')

    ws['A3'] = '№ п/п'
    ws['A3'].style = 'header'

    ws['B3'] = 'Оборудование'
    ws['B3'].style = 'header'

    row = 4

    for contract in contracts:
        cell = ws.cell(row = row, column = 1, value = 'Договор № ' + contract.number + ' от ' + contract.signed.strftime('%d.%m.%Y') + ' (Дата окончания действия договора: ' + contract.enddate.strftime('%d.%m.%Y') + ')')
        cell.style = 'header'
        ws.merge_cells('A' + str(row) + ':B' + str(row))
        row += 1
        for num, equipment in enumerate(contract.eqcontracts.all()):
            cell = ws.cell(row = row, column = 1, value = num + 1)
            cell.style = 'general'
            cell = ws.cell(row = row, column = 2, value = str(equipment))
            cell.style = 'general'
            row += 1

    response = HttpResponse(content = save_virtual_workbook(wb), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=equipments.xlsx'
    return response

@login_required
def AllEquipmentExcelView(request):
    import datetime
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    from openpyxl.styles import NamedStyle, Side, Font, Border, Alignment

    equipments = EquipmentModel.objects.filter(contract__enddate__gt = datetime.datetime.now().date())

    wb = Workbook()
    ws = wb.active
    ws.title = 'Оборудование'

    general = NamedStyle(name = 'general')
    bd = Side(style = 'thin', color = '000000')
    general.font = Font(name = 'Times New Roman', size = 14)
    general.border = Border(left = bd, top = bd, right = bd, bottom = bd)
    general.alignment = Alignment(vertical = 'center', wrap_text = True)
    wb.add_named_style(general)

    header = NamedStyle(name = 'header')
    header.font = Font(name = 'Times New Roman', size = 14, bold = True)
    header.border = Border(left = bd, top = bd, right = bd, bottom = bd)
    header.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
    wb.add_named_style(header)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30

    ws['A1'] = 'Список оборудования'
    ws['A1'].font = Font(name = 'Times New Roman', size = 16,  bold = True)
    ws['A1'].alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
    ws.merge_cells('A1:K1')

    ws['A3'] = 'Номер счета'
    ws['A3'].style = 'header'

    ws['B3'] = 'Номер договора'
    ws['B3'].style = 'header'

    ws['C3'] = 'Производитель'
    ws['C3'].style = 'header'

    ws['D3'] = 'Тип оборудования'
    ws['D3'].style = 'header'

    ws['E3'] = 'Модель оборудования'
    ws['E3'].style = 'header'

    ws['F3'] = 'Серийный номер'
    ws['F3'].style = 'header'

    ws['G3'] = 'Примечание'
    ws['G3'].style = 'header'

    row = 4

    for equipment in equipments:
        cell = ws.cell(row = row, column = 1, value = equipment.account)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 2, value = equipment.contract.number)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 3, value = str(equipment.brand))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 4, value = str(equipment.type or ''))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 5, value = str(equipment.model.name))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 6, value = equipment.sn)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 7, value = equipment.note)
        cell.style = 'general'
        row += 1

    response = HttpResponse(content = save_virtual_workbook(wb), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=equipments.xlsx'
    return response

@login_required
def InstalledBaseExcel(request):
    if not request.user.groups.filter(name = 'Администратор').exists():
        return redirect('login')

    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    from openpyxl.styles import NamedStyle, Side, Font, Border, Alignment

    equipments = EquipmentModel.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Инсталлированная база'

    general = NamedStyle(name = 'general')
    bd = Side(style = 'thin', color = '000000')
    general.font = Font(name = 'Times New Roman', size = 14)
    general.border = Border(left = bd, top = bd, right = bd, bottom = bd)
    general.alignment = Alignment(vertical = 'center', wrap_text = True)
    wb.add_named_style(general)

    header = NamedStyle(name = 'header')
    header.font = Font(name = 'Times New Roman', size = 14, bold = True)
    header.border = Border(left = bd, top = bd, right = bd, bottom = bd)
    header.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
    wb.add_named_style(header)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20

    ws['A1'] = 'Инсталлированная база'
    ws['A1'].font = Font(name = 'Times New Roman', size = 16,  bold = True)
    ws['A1'].alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
    ws.merge_cells('A1:I1')

    ws['A3'] = '№ п/п'
    ws['A3'].style = 'header'

    ws['B3'] = 'Заказчик'
    ws['B3'].style = 'header'

    ws['C3'] = 'Оборудование'
    ws['C3'].style = 'header'

    ws['D3'] = 'Бренд'
    ws['D3'].style = 'header'

    ws['E3'] = 'Тип'
    ws['E3'].style = 'header'

    ws['F3'] = 'Модель'
    ws['F3'].style = 'header'

    ws['G3'] = 'Серийный номер'
    ws['G3'].style = 'header'

    ws['H3'] = 'Старт поддержки'
    ws['H3'].style = 'header'

    ws['I3'] = 'Окончание поддержки'
    ws['I3'].style = 'header'

    row = 4

    for ind, equipment in enumerate(equipments):
        cell = ws.cell(row = row, column = 1, value = ind + 1)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 2, value = str(equipment.contract.client) + ' (' + equipment.contract.client.organization + ')')
        cell.style = 'general'
        cell = ws.cell(row = row, column = 3, value = str(equipment.model) + ' (S/n: ' + str(equipment.sn) + ')')
        cell.style = 'general'
        cell = ws.cell(row = row, column = 4, value = str(equipment.brand))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 5, value = str(equipment.type or ''))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 6, value = str(equipment.model.name))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 7, value = equipment.sn)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 8, value = equipment.contract.signed.strftime('%d.%m.%Y'))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 9, value = equipment.contract.enddate.strftime('%d.%m.%Y'))
        cell.style = 'general'
        row += 1

    response = HttpResponse(content = save_virtual_workbook(wb), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=installed_base.xlsx'
    return response

@login_required
def SpareExcelView(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    from openpyxl.styles import NamedStyle, Side, Font, Border, Alignment

    spares = SpareModel.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'ЗИП'

    general = NamedStyle(name = 'general')
    bd = Side(style = 'thin', color = '000000')
    general.font = Font(name = 'Times New Roman', size = 14)
    general.border = Border(left = bd, top = bd, right = bd, bottom = bd)
    general.alignment = Alignment(vertical = 'center', wrap_text = True)
    wb.add_named_style(general)

    header = NamedStyle(name = 'header')
    header.font = Font(name = 'Times New Roman', size = 14, bold = True)
    header.border = Border(left = bd, top = bd, right = bd, bottom = bd)
    header.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
    wb.add_named_style(header)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 30
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 30
    ws.column_dimensions['O'].width = 30

    ws['A1'] = 'Список ЗИП'
    ws['A1'].font = Font(name = 'Times New Roman', size = 16,  bold = True)
    ws['A1'].alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
    ws.merge_cells('A1:O1')

    ws['A3'] = 'Номер счета'
    ws['A3'].style = 'header'

    ws['B3'] = 'Номер договора'
    ws['B3'].style = 'header'

    ws['C3'] = 'Производитель'
    ws['C3'].style = 'header'

    ws['D3'] = 'Тип запчасти'
    ws['D3'].style = 'header'

    ws['E3'] = 'Наименование оборудования'
    ws['E3'].style = 'header'

    ws['F3'] = 'Наименование запчасти'
    ws['F3'].style = 'header'

    ws['G3'] = 'Парт номер'
    ws['G3'].style = 'header'

    ws['H3'] = 'Серийный номер'
    ws['H3'].style = 'header'

    ws['I3'] = 'Количество'
    ws['I3'].style = 'header'

    ws['J3'] = 'Поставщик'
    ws['J3'].style = 'header'

    ws['K3'] = 'Совместимое оборудование'
    ws['K3'].style = 'header'

    ws['L3'] = 'Наименование склада'
    ws['L3'].style = 'header'

    ws['M3'] = 'Место на складе'
    ws['M3'].style = 'header'

    ws['N3'] = 'Состояние'
    ws['N3'].style = 'header'

    ws['O3'] = 'Примечание'
    ws['O3'].style = 'header'

    row = 4

    for spare in spares:
        cell = ws.cell(row = row, column = 1, value = spare.account)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 2, value = spare.contract)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 3, value = str(spare.brand))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 4, value = str(spare.type or ''))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 5, value = str(spare.equipment or ''))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 6, value = str(spare.model.name))
        cell.style = 'general'

        pn = ''
        for record in spare.pnspare.all():
            pn += record.number + '\n'
        cell = ws.cell(row = row, column = 7, value = pn)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 8, value = spare.sn)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 9, value = spare.quantity)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 10, value = str(spare.provider or ''))
        cell.style = 'general'
        eqs = ''
        for record in spare.cmpspare.all():
            eqs += str(record.equipment) + '\n'
        cell = ws.cell(row = row, column = 11, value = eqs)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 12, value = str(spare.stock or ''))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 13, value = spare.place)
        cell.style = 'general'
        cell = ws.cell(row = row, column = 14, value = str(spare.status or ''))
        cell.style = 'general'
        cell = ws.cell(row = row, column = 15, value = spare.note)
        cell.style = 'general'
        row += 1

    response = HttpResponse(content = save_virtual_workbook(wb), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=spares.xlsx'
    return response

@login_required
def SpareLoadView(request):
    from applications.forms import SpareLoadForm

    if request.method == 'POST':
        form = SpareLoadForm(request.POST, request.FILES)
        if form.is_valid():
            from openpyxl import load_workbook
            from django.http import HttpResponse
            from applications.models import SpareTypeModel, BrandModel, SpareNameModel, ModelModel, StockModel, ProviderModel, EquipmentStatusModel, SpareCompatibleModel, SparePNModel

            wb = load_workbook(filename = request.FILES['excel'].file)
            sheets = wb.sheetnames
            ws = wb[sheets[0]]
            start = 0
            end = ws.max_row
            spares = []

            types = SpareTypeModel.objects.all().values()
            brands = BrandModel.objects.all().values()
            models = SpareNameModel.objects.all().values()
            equipments = ModelModel.objects.all().values()
            stocks = StockModel.objects.all().values()
            providers = ProviderModel.objects.all().values()
            statuses = EquipmentStatusModel.objects.all().values()

            for m in range(1 , end + 1):
                if ws.cell(m, 3).value is not None:
                    start = m + 1
                    break

            for m in range(start , end + 1):
                brand = next((item['id'] for item in brands if item['name'] == ws.cell(m, 3).value), ws.cell(m, 3).value)
                pns = (ws.cell(m, 7).value or '').splitlines()
                eqs = (ws.cell(m, 11).value or '').splitlines()
                for eq in eqs:
                    eq = next((item['id'] for item in equipments if item['name'] == ws.cell(m, 5).value and item['brand'] == brand), ws.cell(m, 5).value)

                spares.append({'id':'', 'type': next((item['id'] for item in types if item['name'] == ws.cell(m, 4).value), ws.cell(m, 4).value),
                    'brand': brand,
                    'model': next((item['id'] for item in models if item['name'] == ws.cell(m, 6).value and item['brand'] == brand), ws.cell(m, 6).value),
                    'sn': ws.cell(m, 8).value or '',
                    'equipment': next((item['id'] for item in equipments if item['name'] == ws.cell(m, 5).value and item['brand'] == brand), ws.cell(m, 5).value),
                    'account': ws.cell(m, 1).value or '',
                    'contract': ws.cell(m, 2).value or '',
                    'stock': next((item['id'] for item in stocks if item['name'] == ws.cell(m, 12).value), ws.cell(m, 12).value),
                    'provider': next((item['id'] for item in providers if item['name'] == ws.cell(m, 10).value), ws.cell(m, 10).value),
                    'place': ws.cell(m, 13).value or '',
                    'status': next((item['id'] for item in statuses if item['name'] == ws.cell(m, 14).value), ws.cell(m, 14).value),
                    'quantity': int(ws.cell(m, 9).value or 0),
                    'note': ws.cell(m, 15).value or '',
                    'pns': pns,
                    'eqs': eqs
                })

            for spare in spares:
                if not str(spare['type']).isdigit():
                    obj, created = SpareTypeModel.objects.get_or_create(name = spare['type'])
                    spare['type'] = obj.id
                if not str(spare['brand']).isdigit():
                    obj, created = BrandModel.objects.get_or_create(name = spare['brand'])
                    spare['brand'] = obj.id
                if not str(spare['model']).isdigit():
                    obj, created = SpareNameModel.objects.get_or_create(name = spare['model'], brand = spare['brand'])
                    spare['model'] = obj.id
                if not str(spare['equipment']).isdigit():
                    obj, created = ModelModel.objects.get_or_create(name = spare['equipment'], brand = spare['brand'])
                    spare['equipment'] = obj.id
                if not str(spare['stock']).isdigit():
                    obj, created = StockModel.objects.get_or_create(name = spare['stock'])
                    spare['stock'] = obj.id
                if not str(spare['provider']).isdigit():
                    obj, created = ProviderModel.objects.get_or_create(name = spare['provider'])
                    spare['provider'] = obj.id
                if not str(spare['status']).isdigit():
                    obj, created = EquipmentStatusModel.objects.get_or_create(name = spare['status'])
                    spare['status'] = obj.id

                obj, created = SpareModel.objects.update_or_create(type = spare['type'], brand = spare['brand'], model = spare['model'],
                    sn = spare['sn'], equipment = spare['equipment'], account = spare['account'], contract = spare['contract'], provider = spare['provider'],
                    status = spare['status'],
                    defaults = {'quantity': spare['quantity'], 'stock': spare['stock'], 'place': spare['place'], 'note': spare['note']},)

                spare['id'] = obj.id

                if spare['eqs']:
                    for eq in spare['eqs']:
                        if not str(eq).isdigit():
                            obj, created = ModelModel.objects.get_or_create(name = eq, brand = spare['brand'])
                            eq = obj.id

                        SpareCompatibleModel.objects.get_or_create(spare = spare['id'], equipment = eq)

                if spare['pns']:
                    for pn in spare['pns']:
                        SparePNModel.objects.get_or_create(spare = spare['id'], number = pn)

            return redirect('/admin/applications/sparemodel/')

    form = SpareLoadForm()
    context = {'form': form}
    return render(request, 'applications/spare_load.html', context)

@login_required
def ReadingEmailSelver(request):
    if not request.user.groups.filter(name = 'Администратор').exists():
        return redirect('login')

    import imaplib
    import email
    import base64
    from email.header import decode_header
    import datetime
    import re
    from applications.models import EmailLastUID, AppCommentModel

    try:
        last_uid = EmailLastUID.objects.get(id = 1)
    except:
        last_uid = None

    date = (datetime.date.today() - datetime.timedelta(1)).strftime("%d-%b-%Y")

    imap = imaplib.IMAP4_SSL(settings.EMAIL_HOST_IMAP)
    imap.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

    imap.select('INBOX', True)

    if last_uid:
        result, data = imap.search('utf-8', 'UID', str(last_uid.uid) + ':*', 'subject', u'заявк'.encode('utf-8'))
    else:
        result, data = imap.search('utf-8', 'SINCE', date, 'subject', u'заявк'.encode('utf-8'))

    email_uids = data[0].split()

    if len(email_uids) > 0:
        latest_email_uid = email_uids[-1]
        first_email_uid = email_uids[0]
        return HttpResponse(latest_email_uid)
        if int(latest_email_uid) < last_uid.uid:
            return HttpResponse('successful')
    else:
        return HttpResponse('successful')

    messages = []
    comments = []
    successful = True

    for message in email_uids:
        result, data = imap.uid('fetch', message, '(RFC822)')
        raw_email = data[0][1].decode('utf-8')

        email_message = email.message_from_string(raw_email)

        try:
            subject = decode_header(email_message['Subject'])[0][0].decode()
        except:
            subject = email_message['Subject']

        app_text = re.search(r"^.*заявк. № \d+", subject, re.IGNORECASE | re.UNICODE)
        if app_text:
            app_number = re.search(r"\d+", app_text.group(), re.IGNORECASE | re.UNICODE).group()
            if app_number:
                from_email = email.utils.parseaddr(email_message['From'])[1]

                if email_message.is_multipart():
                    for payload in email_message.get_payload():
                        body = payload.get_payload(decode = True).decode('utf-8')
                else:
                    body = email_message.get_payload(decode = True).decode('utf-8')

                body_tag_position = body.find('<body')
                if body_tag_position:
                    body = body[body_tag_position:]
                    body_tag_close_position = body.find('>') + 1
                    body = body[body_tag_close_position:]

                messages.append({'text': body, 'application_id': int(app_number), 'author_id': from_email, 'email_date': datetime.datetime.strptime(email_message['Date'], '%a, %d %b %Y %H:%M:%S %z')})


    if messages:
        User = get_user_model()
        users = User.objects.all().values('id', 'email')
        for message in messages:
            message['author_id'] = next((row['id'] for row in users if row['email'] == message['author_id']), None)
            if message['author_id']:
                try:
                    AppCommentModel.objects.create(**message)
                except:
                    successful = False

    if successful:
        if last_uid:
            last_uid.success = True
            last_uid.uid = int(latest_email_uid) + 1
            last_uid.pubdate = datetime.datetime.now()
            last_uid.save()
        else:
            EmailLastUID.objects.create(success = True, uid = latest_email_uid)
    return HttpResponse(successful)

@login_required
def TestTaskView(request):
    from applications.tasks import CloseApplication

    rst = CloseApplication.apply().get()
    return HttpResponse(rst)

@login_required
def TestView(request):
    formset = AppDocumentsFormset(instance = GetApplication(160555))
    return render(request, 'applications/test.html', {'formset': formset})


@login_required
def ListApplicationView(request): #Список заявок
    context = ApplicationsListAPIView().as_view()(request = request, page = 1).data
    return render(request, 'applications/list.html', context)

@login_required
@permission_required('applications.add_applicationmodel')
def NewAddApplicationView(request): #Создание заявки
    permissions = {
        'is_admin': request.user.groups.filter(name = 'Администратор').exists(),
        'is_engineer': request.user.groups.filter(name = 'Инженер').exists(),
        'is_staff': is_admin_or_engineer(request.user)
    }

    if not permissions['is_staff']:
        form = ClientApplicationForm()
    else:
        form = StaffApplicationForm()

    context = {'form': form, 'permissions': permissions}
    return render(request, 'applications/new_add.html', context)
