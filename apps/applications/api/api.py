import math
from io import BytesIO

from datetime import datetime
from dateutil import tz

from itertools import chain

from django.contrib.auth import get_user_model
from django.urls import reverse
from django.http import FileResponse, HttpResponse
from django.conf import settings

from django.db.models import Subquery, OuterRef, Value, Q, F, Func, Case, When, IntegerField, CharField, Prefetch, Sum
from django.db.models.functions import Concat, Trim

from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import APIException
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FileUploadParser, FormParser

from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiExample, OpenApiResponse

from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import NamedStyle, Side, Font, Border, Alignment

from applications.models import *
from contracts.models import ContractModel, ContractEquipmentModel
# from accounts.models import OrganizationContactModel
from spares.models import SpareModel

from spares.api.api import SparesListAPIView

from accounts.api.api import ContactsListAPIView

from integrator.apps.api_services import EngineersListAPIView
from applications.api.serializers import *

from integrator.apps.functions import is_admin_or_engineer, get_prms_from_ids, send_email, send_telegram

@extend_schema(
    tags = ['Список приоритетов заявки (Done)'],
    summary = 'Список приоритетов заявки',
    description = '<ol><li>"id" - Идентификатор</li><li>"name" - Наименование</li><li>"priority" - Приоритет отображения (упорядочивание от меньшего к большему)</li></ol>'
)
class AppPrioritiesListAPIView(generics.ListAPIView):
    queryset = AppPriorityModel.objects.all()
    serializer_class = AppPrioritySerializer
    permission_classes = [IsAuthenticated]

@extend_schema(
    tags = ['Список статусов заявки (Done)'],
    summary = 'Список статусов заявки',
    description = '<ol><li>"id" - Идентификатор</li><li>"name" - Наименование</li><li>"priority" - Приоритет отображения (упорядочивание от меньшего к большему)</li></ol>'
)
class AppStatusesAPIView(generics.ListAPIView):
    queryset = StatusModel.objects.all()
    serializer_class = AppStatusSerializer
    permission_classes = [IsAuthenticated]

class ApplicationListPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 100

class ApplicationsListAPIView(APIView):
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Список заявок',
        description = '<b>Внимание!</b> Список заявок отображается в зависимости от прав текущего пользователя\
            <p>"applications" - список заявок, где</p>\
            <ol><li>"id" - Идентификатор / номер заявки</li><li>"equipment_id" - Идентификатор оборудования</li>\
            <li>"equipment_name" - Наименование оборудования</li><li>"status_id" - Идентификатор статуса заявки</li>\
            <li>"status_name" - Наименование статуса заявки</li><li>"problem" - Описание проблемы</li>\
            <li>"priority_id" - Идентификатор приоритета заявки <b>(Доступен только персоналу)</b></li>\
            <li>"priority_name" - Наименование приоритета заявки <b>(Доступен только персоналу)</b></li>\
            <li>"engineer_id" - Идентификатор назначенного инженера <b>(Доступен только персоналу)</b></li>\
            <li>"engineer_last_name" - Фамилия назначенного инженера <b>(Доступен только персоналу)</b></li>\
            <li>"engineer_name" - Полное имя назначенного инженера, если указано, иначе - e-mail <b>(Доступен только персоналу)</b></li>\
            <li>"organization_id" - Идентификатор организации заказчика <b>(Доступен только персоналу)</b></li>\
            <li>"organization_name" - Наименование организации заказчика <b>(Доступен только персоналу)</b></li>\
            <li>"end_user_organization_id" - Идентификатор организации конечного пользователя <b>(Доступен только персоналу)</b></li>\
            <li>"end_user_organization_name" - Наименование организации конечного пользователя <b>(Доступен только персоналу)</b></li>\
            <li>"formatted_date" - Дата создания</li></ol>\
            <p>"statuses" - Список статусов заявок для фильтра, где</p>\
            <ol><li>"id" - Идентификатор статуса</li><li>"name" - Наименование статуса</li><li>"priority" - Приоритет отображения статуса (упорядочивание от меньшего к большему)</li></ol>\
            <p>"permissions" - список условных разрешений пользователя, где</p>\
            <ol><li>"is_admin" - true, если пользователь - администратор</li><li>"is_engineer" - true, если пользователь - инженер</li><li>"is_staff" - true, если пользователь - персонал (администратор или инженер)</li></ol>\
            <p>"application_link" - тип ссылки для заявок (нужен для текущего приложения)</p>\
            <p>"pagination" - пагинация (используется при разбиении списка на страницы), где</p>\
            <ol><li>"page" - Номер страницы</li><li>"num_pages" - Количество страниц</li><li>"pages_range" - Диапазон страниц (нужен для текущего приложения)</li>\
            <li>"has_next" - true, если есть следующая страница (нужен для текущего приложения)</li><li>"has_previous" - true, если есть предыдущая страница (нужен для текущего приложения)</li></ol>\
            <p>"priorities" - Список приоритетов заявки для фильтра <b>(Доступен только персоналу)</b>, где</p>\
            <ol><li>"id" - Идентификатор приоритета</li><li>"name" - Наименование приоритета</li><li>"priority" - Приоритет отображения (упорядочивание от меньшего к большему)</li></ol>\
            <p>"engineers" - Список инженеров для фильтра <b>(Доступен только персоналу)</b>, где</p>\
            <ol><li>"id" - Идентификатор инженера</li><li>"name" - Полное имя инженера, если указано, иначе - e-mail</li></ol>\
            <p>"client" - Данные заказчика <b>(Доступен только заказчику)</b>, где</p>\
            <ol><li>"id" - Идентификатор заказчика</li><li>"fio" - Полное имя заказчика</li><li>"inn" - ИНН заказчика</li>\
            <li>"address" - Адрес заказчика</li><li>"email" - E-mail заказчика</li><li>"phone" - Телефон заказчика</li>\
            <li>"organization_id" - Идентификатор организации заказчика</li><li>"organization_name" - Наименование организации заказчика</li>\
            <li>"contracts" - Список договоров заказчика, где</li>\
            <ul><li>"number" - Номер договора</li><li>"formatted_signed" - Начало договора</li><li>"formatted_enddate" - Окончание договора</li>\
            <li>"eqcontracts" - Список оборудования договора, где\
            <ul><li>"name" - Наименование оборудования</li><li>"sn" - Серийный номер оборудования</li></ul></li></ul></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = ListSerializer())},
        parameters = [
            OpenApiParameter(name = 'number', description = 'Номер / часть номера заявки', required = False, type = int),
            OpenApiParameter(name = 'period', description = 'Период создания заявки', required = False, type = str,
                examples = [OpenApiExample('Пример', value = '01.10.2023 - 31.12.2024'),]
            ),
            OpenApiParameter(name = 'status_ids', description = 'Идентификаторы статусов заявок через разделитель "|"', required = False, type = str,
                examples = [
                    OpenApiExample('Пример #1', value = '1|2|6|'),
                    OpenApiExample('Пример #2', value = 'all|'),
                ]
            ),
            OpenApiParameter(name = 'priority_ids', description = 'Идентификаторы приоритетов заявок через разделитель "|" (для инженеров и администраторов)', required = False, type = str,
                examples = [
                    OpenApiExample('Пример #1', value = '1|2|'),
                    OpenApiExample('Пример #2', value = 'all|'),
                ]
            ),
            OpenApiParameter(name = 'organization', description = 'Наименование / часть наименования организации (для инженеров и администраторов)', required = False, type = str),
            OpenApiParameter(name = 'engineer_ids', description = 'Идентификаторы инженеров через разделитель "|" (для инженеров и администраторов)', required = False, type = str,
                examples = [
                    OpenApiExample('Пример #1', value = '10|2|'),
                    OpenApiExample('Пример #2', value = 'all|'),
                ]
            ),
            OpenApiParameter(name = 'mine', description = 'Заявки, назначенные мне (для инженера)', required = False, type = int,
                examples = [
                    OpenApiExample('Пример #1', value = 1),
                    OpenApiExample('Пример #2', value = 0),
                ]
            ),
            OpenApiParameter(name = 'opened', description = 'Исключить закрытые заявки (для инженеров и администраторов)', required = False, type = int,
                examples = [
                    OpenApiExample('Пример #1', value = 1),
                    OpenApiExample('Пример #2', value = 0),
                ]
            ),
            OpenApiParameter(name = 'equipment', description = 'Наименование / часть наименования / серийный номер / часть серийного номера оборудования', required = False, type = str),
            OpenApiParameter(name = 'page', description = 'Номер страницы', required = False, type = int),
        ],
    )
    def get(self, request, *args, **kwargs):
        User = get_user_model()
        prms = {}
        pagination = {}
        data = []
        page = None
        permissions = {
            'is_admin': request.user.groups.filter(name = 'Администратор').exists(),
            'is_engineer': request.user.groups.filter(name = 'Инженер').exists(),
            'is_staff': is_admin_or_engineer(request.user)
        }

        application_link = 'edit-application' if permissions['is_staff'] else 'app-details'
        if not permissions['is_staff']:
            client = User.objects.filter(id = request.user.id)\
                .annotate(
                    fio = Trim(Concat('last_name', Value(' '), 'first_name')),
                    organization_name = F('organization__name')
                )\
                .values('id', 'fio', 'inn', 'address', 'email', 'phone', 'organization_id', 'organization_name')[0]

            client['contracts'] = GetClientsContractsAPIView().get(request).data

            client = ClientSerializer(client).data
            prms['client__organization_id'] = client['organization_id']

        query_params_keys = request.query_params.keys()

        if 'number' in query_params_keys:
            prms['id__icontains'] = request.query_params.get('number')
        if 'period' in query_params_keys:
            dates = request.query_params.get('period').split(' - ')
            prms['pubdate__gte'] = datetime.strptime(dates[0], '%d.%m.%Y')
            prms['pubdate__lte'] = datetime.strptime(dates[1], '%d.%m.%Y')
        if 'status_ids' in query_params_keys:
            prms.update(get_prms_from_ids(request.query_params.get('status_ids'), 'status'))
        if 'priority_ids' in query_params_keys:
            prms.update(get_prms_from_ids(request.query_params.get('priority_ids'), 'priority'))
        if 'organization' in query_params_keys:
            prms['client_id__in'] = User.objects.filter(organization__name__icontains = request.query_params.get('organization')).values('id').distinct()
        if 'engineer_ids' in query_params_keys:
            prms.update(get_prms_from_ids(request.query_params.get('engineer_ids'), 'engineer'))
        if 'mine' in query_params_keys and int(request.query_params.get('mine')):
            prms['engineer'] = request.user
        if 'opened' in query_params_keys and int(request.query_params.get('opened')):
            prms['status_id__in'] = [1, 2, 3, 6]
        if 'equipment' in query_params_keys:
            equipment = request.query_params.get('equipment')
            prms['equipment_id__in'] = ContractEquipmentModel.objects.filter(
                Q(equipment__model__name__icontains = equipment) | Q(sn__icontains = equipment)
            ).values('id')
        if 'page' in query_params_keys:
            page = int(request.query_params.get('page'))
        if 'page' in kwargs:
            page = int(kwargs['page'])

        fields = ['id', 'formatted_date', 'equipment_id', 'equipment_name', 'status_id', 'status_name', 'problem',]

        if permissions['is_staff']:
            fields.extend([
                'priority_id', 'priority_name', 'organization_id', 'organization_name',
                'end_user_organization_id', 'end_user_organization_name', 'engineer_id', 'engineer_name',
                'engineer_last_name'
            ])

        applications = ApplicationModel.objects.filter(**prms)\
            .annotate(
                organization = Subquery(User.objects.filter(id = OuterRef('client_id')).values('organization__name')),
                equipment_name = Concat('equipment__equipment__brand__name', Value(' '), 'equipment__equipment__model__name', Value(' (S/n:'), 'equipment__sn', Value(')')),
                status_name = F('status__name'),
                priority_name = Func(F('priority__name'), Value(''), function = 'IFNULL', output_field = CharField()),
                organization_id = F('client__organization_id'),
                organization_name = F('client__organization__name'),
                end_user_organization_id = F('equipment__contract__end_user__organization__id'),
                end_user_organization_name = Func(F('equipment__contract__end_user__organization__name'), Value(''), function = 'IFNULL', output_field = CharField()),
                engineer_last_name = Trim('engineer__last_name'),
                engineer_name = Case(
                    When(
                        Q(engineer__isnull = False) & Q(engineer__last_name__isnull = False) & ~Q(engineer_last_name = ''),
                        then = Concat('engineer__first_name', Value(' '), 'engineer__last_name')
                    ),
                    When(engineer_last_name = '', then = F('engineer__email')),
                    default = Value(''), output_field = CharField()),
                formatted_date = Func(
                    Func(F('pubdate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                    Value('%d.%m.%Y %H:%i'), function = 'DATE_FORMAT', output_field = CharField()
                )
            )

        applications = applications.values(*fields)

        if page:
            paginator = ApplicationListPagination()
            applications = paginator.paginate_queryset(applications, request)

            num_pages = math.ceil(paginator.page.paginator.count / paginator.page_size)
            pagination = {
                'page': page, 'num_pages': num_pages, 'pages_range': list(range(1, num_pages + 1)),
                'has_next': True if paginator.get_next_link() else False,
                'has_previous': True if paginator.get_previous_link() else False,
            }

        statuses = AppStatusesAPIView.as_view()(request._request).data

        data = {
            'applications': applications,
            'statuses': statuses,
            'permissions': permissions,
            'application_link': application_link,
            'pagination': pagination
        }

        if permissions['is_staff']:
            data['priorities'] = AppPrioritiesListAPIView.as_view()(request._request).data
            data['engineers'] = EngineersListAPIView.as_view()(request._request).data
        else:
            data['client'] = client

        return Response(data, status=status.HTTP_200_OK)

class ApplicationsExcelAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Экспорт списка заявок (доступно только администраторам)',
        description = 'Response - файл с данными',
        parameters=[
            OpenApiParameter(name = 'number', description = 'Номер / часть номера заявки', required = False, type = int),
            OpenApiParameter(name = 'period', description = 'Период создания заявки', required = False, type = str,
                examples = [OpenApiExample('Пример', value = '01.10.2023 - 31.12.2024'),]
            ),
            OpenApiParameter(name = 'status_ids', description = 'Идентификаторы статусов заявок через разделитель "|"', required = False, type = str,
                examples = [
                    OpenApiExample('Пример #1', value = '1|2|6|'),
                    OpenApiExample('Пример #2', value = 'all|'),
                ]
            ),
            OpenApiParameter(name = 'priority_ids', description = 'Идентификаторы приоритетов заявок через разделитель "|" (для инженеров и администраторов)', required = False, type = str,
                examples = [
                    OpenApiExample('Пример #1', value = '1|2|'),
                    OpenApiExample('Пример #2', value = 'all|'),
                ]
            ),
            OpenApiParameter(name = 'organization', description = 'Наименование / часть наименования организации (для инженеров и администраторов)', required = False, type = str),
            OpenApiParameter(name = 'engineer_ids', description = 'Идентификаторы инженеров через разделитель "|" (для инженеров и администраторов)', required = False, type = str,
                examples = [
                    OpenApiExample('Пример #1', value = '10|2|'),
                    OpenApiExample('Пример #2', value = 'all|'),
                ]
            ),
            OpenApiParameter(name = 'mine', description = 'Заявки, назначенные мне (для инженера)', required = False, type = int,
                examples = [
                    OpenApiExample('Пример #1', value = 1),
                    OpenApiExample('Пример #2', value = 0),
                ]
            ),
            OpenApiParameter(name = 'opened', description = 'Исключить закрытые заявки (для инженеров и администраторов)', required = False, type = int,
                examples = [
                    OpenApiExample('Пример #1', value = 1),
                    OpenApiExample('Пример #2', value = 0),
                ]
            ),
            OpenApiParameter(name = 'equipment', description = 'Наименование / часть наименования / серийный номер / часть серийного номера оборудования', required = False, type = str),
            OpenApiParameter(name = 'page', description = 'Номер страницы', required = False, type = int),
        ],
    )
    def get(self, request, *args, **kwargs):
        data = ApplicationsListAPIView.as_view()(request._request).data

        wb = Workbook()
        with NamedTemporaryFile(delete = True) as tmp:

            ws = wb.active
            ws.title = 'Список заявок'

            bd = Side(style = 'thin', color = '000000')

            header = NamedStyle(name = 'header')
            header.font = Font(name = 'Times New Roman', size = 14, bold = True)
            header.border = Border(left = bd, top = bd, right = bd, bottom = bd)
            header.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
            wb.add_named_style(header)

            general = NamedStyle(name = 'general')
            general.font = Font(name = 'Times New Roman', size = 14)
            general.border = Border(left = bd, top = bd, right = bd, bottom = bd)
            general.alignment = Alignment(vertical = 'center', wrap_text = True)
            wb.add_named_style(general)

            ws['A1'] = 'Список заявок'
            ws['A1'].font = Font(name = 'Times New Roman', size = 16,  bold = True)
            ws['A1'].alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
            ws.merge_cells('A1:I1')

            columns = [
                {'name': 'Идентификатор', 'width': 15},
                {'name': 'Дата создания', 'width': 15},
                {'name': 'Приоритет заявки', 'width': 15},
                {'name': 'Организация', 'width': 20},
                {'name': 'Конечный пользователь', 'width': 20},
                {'name': 'Оборудование', 'width': 30},
                {'name': 'Статус заявки', 'width': 15},
                {'name': 'Краткое описание', 'width': 45},
                {'name': 'Назначенный инженер', 'width': 20},
            ]

            column_id = 1

            for column in columns:
                column_letter = get_column_letter(column_id)
                ws[column_letter + '3'] = column['name']
                ws[column_letter + '3'].style = 'header'
                ws.column_dimensions[column_letter].width = column['width']
                column_id += 1

            row_id = 4
            for application in data['applications']:
                ws['A' + str(row_id)] = application['id']
                ws['B' + str(row_id)] = application['formatted_date']#.astimezone(tz.gettz(settings.TIME_ZONE)).strftime('%d.%m.%Y %H:%M')
                ws['C' + str(row_id)] = application['priority_name']
                ws['D' + str(row_id)] = application['organization_name']
                ws['E' + str(row_id)] = application['end_user_organization_name']
                ws['F' + str(row_id)] = application['equipment_name']
                ws['G' + str(row_id)] = application['status_name']
                ws['H' + str(row_id)] = application['problem']
                ws['I' + str(row_id)] = application['engineer_name']

                column_id = 1
                for column in columns:
                    ws[get_column_letter(column_id) + str(row_id)].style = 'general'
                    column_id += 1

                row_id += 1

            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content = stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=applications_list.xlsx'
        return response

class GetClientsContractsAPIView(APIView): #Список договоров клиента с оборудованием
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        tags = ['Личный кабинет клиента (Функционал доступен только клиентам) (Done)'],
        summary = 'Список договоров с оборудованием',
        description = '<ol><li>"number" - Номер договора</li><li>"formatted_signed" - Начало договора</li><li>"formatted_enddate" - Окончание договора</li>\
        <li>"eqcontracts" - Список оборудования договора, где\
        <ul><li>"name" - Наименование оборудования</li><li>"sn" - Серийный номер оборудования</li></ul></li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = ContractSerializer(many = True))},
    )
    def get(self, request, *args, **kwargs):

        if is_admin_or_engineer(request.user):
            return Response({'message': 'Функционал доступен только клиентам'}, status = status.HTTP_403_FORBIDDEN)

        contracts = ContractModel.objects.filter(client__organization_id = request.user.organization_id)\
            .annotate(formatted_signed = Func(
                    Func(F('signed'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                    Value('%d.%m.%Y'), function = 'DATE_FORMAT', output_field = CharField()
                ),
                formatted_enddate = Func(
                    Func(F('enddate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                    Value('%d.%m.%Y'), function = 'DATE_FORMAT', output_field = CharField()
                )
            )\
            .prefetch_related(Prefetch('eqcontracts', queryset = ContractEquipmentModel.objects.annotate(
                name = Concat('equipment__brand__name', Value(' '), 'equipment__model__name', Value(' (S/n: '), 'sn' , Value(')'))
            )))

        serializer = ContractSerializer(contracts, many = True)

        return Response(serializer.data, status = status.HTTP_200_OK)

class ClientsContractsExcelAPIView(APIView): #Список договоров клиента с оборудованием. Выгрузка в Excel
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        tags = ['Личный кабинет клиента (Функционал доступен только клиентам) (Done)'],
        summary = 'Выгрузка в Excel списка договоров с оборудованием',
        description = 'Response - файл с данными'
    )
    def get(self, request, *args, **kwargs):
        if is_admin_or_engineer(request.user):
            return Response({'message': 'Функционал доступен только клиентам'}, status = status.HTTP_403_FORBIDDEN)

        contracts = GetClientsContractsAPIView().get(request).data

        wb = Workbook()
        with NamedTemporaryFile(delete = True) as tmp:

            ws = wb.active
            ws.title = 'Оборудование'

            bd = Side(style = 'thin', color = '000000')

            header = NamedStyle(name = 'header')
            header.font = Font(name = 'Times New Roman', size = 14, bold = True)
            header.border = Border(left = bd, top = bd, right = bd, bottom = bd)
            header.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
            wb.add_named_style(header)

            general = NamedStyle(name = 'general')
            general.font = Font(name = 'Times New Roman', size = 14)
            general.border = Border(left = bd, top = bd, right = bd, bottom = bd)
            general.alignment = Alignment(vertical = 'center', wrap_text = True)
            wb.add_named_style(general)

            ws['A1'] = 'Список оборудования'
            ws['A1'].font = Font(name = 'Times New Roman', size = 16,  bold = True)
            ws['A1'].alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
            ws.merge_cells('A1:B1')

            columns = [
                {'name': '№ п/п', 'width': 10},
                {'name': 'Оборудование', 'width': 90},
            ]

            column_id = 1

            for column in columns:
                column_letter = get_column_letter(column_id)
                ws[column_letter + '3'] = column['name']
                ws[column_letter + '3'].style = 'header'
                ws.column_dimensions[column_letter].width = column['width']
                column_id += 1

            row_id = 4
            for contract in contracts:
                cell = ws.cell(row = row_id, column = 1, value = 'Договор № ' + contract['number'] + ' от ' + contract['formatted_signed'] + ' (Дата окончания действия договора: ' + contract['formatted_enddate'] + ')')
                cell.style = 'header'
                ws.merge_cells('A' + str(row_id) + ':B' + str(row_id))
                row_id += 1

                for num, equipment in enumerate(contract['eqcontracts']):
                    cell = ws.cell(row = row_id, column = 1, value = num + 1)
                    cell.style = 'general'
                    cell = ws.cell(row = row_id, column = 2, value = equipment['name'])
                    cell.style = 'general'
                    row_id += 1

            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content = stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=equipments.xlsx'
        return response

class GetClientsContactsAPIView(APIView): #Создание заявки. Список контактных лиц и доступного оборудования
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Создание заявки. Список контактных лиц и доступного оборудования',
        description = '<p>"contacts" - Список контактных лиц организации, где</p><ol><li>"id" - Идентификатор контакта</li><li>"fio" - ФИО контакта</li></ol>\
        <p>"equipments" - Список доступного оборудования организации, где</p><ol><li>"id" - Идентификатор оборудования в договоре</li>\
        <li>"equipment_id" - Идентификатор оборудования</li><li>"equipment_name" - Наименование оборудования</li></ol>',
        parameters = [
            OpenApiParameter(name = 'client_id', description = 'Идентификатор заказчика', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = ContactsEquipmentSerializer())},
    )
    def get(self, request, client_id, *args, **kwargs):
        is_staff = is_admin_or_engineer(request.user)
        if not is_staff:
            client_id = request.user.id

        User = get_user_model()

        try:
            client = User.objects.get(id = client_id)
        except User.DoesNotExist:
            return Response([{'error': 'Указанный клиент не найден'}], status=status.HTTP_406_NOT_ACCEPTABLE)

        contacts = ContactsListAPIView().get(request, client.organization_id).data

        equipments = ContractEquipmentModel.objects.filter(
            Q(contract__client__organization_id = client.organization_id)
            & Q(contract__enddate__gt = datetime.now().date())
        ).annotate(
            equipment_name = Concat('equipment__brand__name', Value(' '), 'equipment__model__name', Value(' (S/n:'), 'sn', Value(')')),
        ).values('id', 'equipment_name', 'equipment_id')

        return Response({'contacts': contacts, 'equipments': equipments}, status=status.HTTP_200_OK)

class GetClientsEquipmentsAPIView(APIView): #Список доступного для клиента оборудования
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        tags = ['Редактирование заявки (Done)'],
        summary = 'Список доступного для клиента оборудования',
        description = '<p>"equipments" - Список доступного для клиента оборудования, где</p><ol><li>"id" - Идентификатор оборудования в договоре</li>\
        <li>"equipment_id" - Идентификатор оборудования</li><li>"equipment_name" - Наименование оборудования</li></ol>',
        parameters = [
            OpenApiParameter(name = 'client_id', description = 'Идентификатор клиента', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = OrganizationEquipmentSerializer(many = True))},
    )
    def get(self, request, client_id, *args, **kwargs):
        User = get_user_model()

        try:
            client = User.objects.get(id = client_id)
        except User.DoesNotExist:
            return Response([{'error': 'Указанный клиент не найден'}], status=status.HTTP_406_NOT_ACCEPTABLE)

        equipments = ContractEquipmentModel.objects.filter(
            Q(contract__client__organization_id = client.organization_id)
            & Q(contract__enddate__gt = datetime.now().date())
        ).annotate(
            equipment_name = Concat('equipment__brand__name', Value(' '), 'equipment__model__name', Value(' (S/n:'), 'sn', Value(')')),
        ).values('id', 'equipment_name', 'equipment_id')

        return Response({'equipments': equipments}, status=status.HTTP_200_OK)

class AddApplicationAPIView(APIView): #Создание заявки
    permission_classes = [IsAuthenticated,]
    parser_classes = [MultiPartParser,]

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Создание заявки. Реквизиты формы (Для текущего приложения)',
        description = '<b>Внимание!</b> Список заказчиков "clients" доступен только для персонала\
        <p>"priorities" - Список приоритетов заявки, где</p>\
        <ol><li>"id" - Идентификатор</li><li>"name" - Наименование</li><li>"priority" - Приоритет отображения (упорядочивание от меньшего к большему)</li></ol>\
        <p>"permissions" - список условных разрешений пользователя, где</p>\
        <ol><li>"is_admin" - true, если пользователь - администратор</li><li>"is_engineer" - true, если пользователь - инженер</li><li>"is_staff" - true, если пользователь - персонал (администратор или инженер)</li></ol>\
        <p>"clients" - список заказчиков, где</p>\
        <ol><li>"id" - Идентификатор заказчика</li><li>"organization_name" - Наименование организации заказчика</li><li>"name" - ФИО заказчика, если указано, иначе - e-mail</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = AppFormSerializer())},
    )
    def get(self, request, *args, **kwargs):
        User = get_user_model()

        permissions = {
            'is_admin': request.user.groups.filter(name = 'Администратор').exists(),
            'is_engineer': request.user.groups.filter(name = 'Инженер').exists(),
            'is_staff': is_admin_or_engineer(request.user)
        }

        priorities = AppPrioritiesListAPIView.as_view()(request._request).data

        data = {'priorities': priorities, 'permissions': permissions}

        if permissions['is_staff']:
            clients = User.objects.filter(groups__name = 'Заказчик')\
                .annotate(organization_name = F('organization__name'),
                    name = Case(
                        When(
                            Q(last_name__isnull = False) & ~Q(last_name = ''),
                            then = Concat('first_name', Value(' '), 'last_name')
                        ),
                        When(last_name = '', then = F('email')),
                        default = Value(''), output_field = CharField())
                    )\
                .values('id', 'organization_name', 'name')
            data['clients'] = clients

        return Response(data, status = status.HTTP_200_OK)

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Создание заявки',
        description = '<p><b>Внимание!</b> Если текущий пользователь - не персонал, то в поле "client" автоматически встает его идентификатор</p>\
        <p><b>Внимание!</b> Если текущий пользователь - инженер, то в поле "engineer" автоматически встает его идентификатор, а в поле "status" - идентификатор статуса "В работе", иначе "status" - идентификатор статуса "Зарегистрирована"</p>\
        <ol><li>"priority" - Идентификатор приоритета заявки</li><li>"problem" - Описание проблемы</li><li>"contact" - Идентификатор контактного лица</li>\
        <li>"client" - Идентификатор заказчика</li><li>"equipment" - Идентификатор оборудования</li><li>"engineer" - Идентификатор инженера <b>(устанавливается автоматически, передавать не надо)</b></li>\
        <li>"status" - Идентификатор статуса заявки <b>(устанавливается автоматически, передавать не надо)</b></li>\
        <li>"documents" - Список прикрепленных файлов, где<ul><li>"document" - файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b></li><li>"name" - Наименование файла (если не указано, то значение будет равно имени файла)</li></ul></li></ol>',
        request = {'application/json': AddApplicationSerializer()},
        responses = {(201, 'application/json'): OpenApiResponse(response = AddApplicationSerializer())}
    )
    def post(self, request, *args, **kwargs):
        Users = get_user_model()

        if not is_admin_or_engineer(request.user):
            request.data._mutable = True
            request.data['client'] = request.user.id

        if not request.data['client']:
            return Response({'message': 'Неверно указан заказчик'}, status = status.HTTP_400_BAD_REQUEST)

        if request.user.groups.filter(name = 'Инженер').exists():
            request.data['engineer'] = request.user.id
            request.data['status'] = 6
        else:
            request.data['status'] = 1

        equipment_exists = ContractEquipmentModel.objects.filter(Q(id = request.data['equipment'])
            & Q(contract__client__organization_id = Users.objects.get(id = request.data['client']).organization_id)
            & Q(contract__enddate__gt = datetime.now().date())
        ).exists()

        if not equipment_exists:
            return Response({'message': 'Оборудование выбрано неверно'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = AddApplicationSerializer(data = request.data, context = {'request': request})

        if serializer.is_valid():
            application = serializer.save()
            AppStatusModel.objects.create(status = application.status, application = application)
            AppHistoryModel.objects.create(type = 1, text = 'Заявка создана, присвоен статус "' + str(application.status) + '"', application = application, author = request.user)

            contact_email = application.contact.email
            params = {
                'id': application.id,
                'url': request.build_absolute_uri(application.get_absolute_url()),
                'type': 'add'
            }
            if send_email(params = params, title = 'Создание заявки № ' + str(application.id), send_to = [contact_email]):
                AppHistoryModel.objects.create(type = 3, text = 'Отправлено сообщение о создании заявки на адрес электронной почты ' + contact_email, application = application, author = request.user)
            else:
                AppHistoryModel.objects.create(type = 3, text = 'Не удалось отправить сообщение о создании заявки на адрес электронной почты ' + contact_email, application = application, author = request.user)

            if send_telegram(params = params):
                AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение о создании заявки в телеграм-канал', application = application, author = request.user)
            else:
                AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение о создании заявки в телеграм-канал', application = application, author = request.user)

            if application.status_id == 6:
                contact_email = application.engineer.email
                if send_email(params = {
                        'id': application.id,
                        'url': request.build_absolute_uri(application.get_absolute_url()),
                        'status': application.engineer,
                        'type': 'engineer'
                    }, title = 'Назначен инженер', send_to = [contact_email]):
                    AppHistoryModel.objects.create(type = 4, text = 'Отправлено сообщение о назначении инженера "' + str(application.engineer) + '" на адрес электронной почты ' + contact_email, application = application, author = request.user)
                else:
                    AppHistoryModel.objects.create(type = 4, text = 'Не удалось отправить сообщение о назначении инженера "' + str(application.engineer) + '" на адрес электронной почты ' + contact_email, application = application, author = request.user)

            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class EditApplicationAPIView(APIView): #Редактирование заявки
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Редактирование заявки. Страница редактирования заявки',
        description = '<p>"application" - Данные заявки, где</p><ol><li>"id" - Идентификатор / номер заявки</li><li>"formatted_date" - Дата создания</li>\
        <li>"status_id" - Идентификатор статуса заявки</li><li>"status_name" - Наименование статуса заявки</li>\
        <li>"priority_id" - Идентификатор приоритета заявки</li><li>"priority_name" - Наименование приоритета заявки</li>\
        <li>"engineer_id" - Идентификатор инженера</li><li>"engineer_name" - ФИО инженера, если указано, иначе e-mail</li>\
        <li>"contact_id" - Идентификатор контактного лица</li><li>"contact_name" - ФИО контактного лица</li><li>"contact_email" - e-mail контактного лица</li>\
        <li>"contact_phone" - Телефон контактного лица</li><li>"support_level" - Наименование уровня поддержки</li>\
        <li>"vendor_name" - Наименование вендора</li><li>"equipment_id" - Идентификатор оборудования</li><li>"equipment_name" - Наименование оборудования</li>\
        <li>"changed" - True, если оборудование в заявке было заменено</li><li>"problem" - Описание проблемы</li>\
        <li>"documents" - Список прикрепленных файлов, где<ul><li>"document" - файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b></li><li>"name" - Наименование файла (если не указано, то значение будет равно имени файла)</li></ul></li></ol>\
        <p>"history" - Действия заявки, где</p><ol><li>"id" - Идентификатор комментария / события</li><li>"pubdate" - Дата создания</li>\
        <li>"text" - Текст</li><li>"author_name" - ФИО автора, если указано, иначе e-mail</li><li>"formatted_date" - Форматированная дата создания</li>\
        <li>"record_type" - Тип записи ("comment" - комментарий, "status" - событие)</li><li>"hide" - True, если запись - комментарий, и его нужно скрыть от пользователя</li></ol>\
        <p>"spares" - Список ЗИП, где</p><ol><li>"id" - Идентификатор ЗИП</li><li>"spare" - Отображение ЗИП</li><li>"name" - Наименование ЗИП</li>\
        <li>"sn" - Серийный номер ЗИП</li><li>"description" - Описание ЗИП</li><li>"barcode" - Ссылка на штрих-код</li></ol>\
        <p>"permissions" - список условных разрешений пользователя, где</p>\
        <ol><li>"is_admin" - true, если пользователь - администратор</li><li>"is_engineer" - true, если пользователь - инженер</li><li>"is_staff" - true, если пользователь - персонал (администратор или инженер)</li></ol>',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = EditAppFormSerializer())}
    )
    def get(self, request, application_id, *args, **kwargs):
        permissions = {
            'is_admin': request.user.groups.filter(name = 'Администратор').exists(),
            'is_engineer': request.user.groups.filter(name = 'Инженер').exists(),
            'is_staff': is_admin_or_engineer(request.user)
        }

        application = ApplicationModel.objects.annotate(
            formatted_date = Func(
                Func(F('pubdate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                Value('%d.%m.%Y %H:%i'), function = 'DATE_FORMAT', output_field = CharField()
            ),
            status_name = F('status__name'),
            priority_name = F('priority__name'),
            engineer_name = Trim(Case(
                When(
                    Q(engineer__isnull = False) & Q(engineer__last_name__isnull = False),
                    then = Concat('engineer__first_name', Value(' '), 'engineer__last_name')
                ),
                When(engineer__last_name = '', then = F('engineer__email')),
                default = Value(''), output_field = CharField()
            )),
            contact_name = F('contact__fio'),
            contact_email = F('contact__email'),
            contact_phone = F('contact__phone'),
            support_level = F('equipment__support__name'),
            vendor_name = Case(
                When(
                    Q(equipment__equipment__vendor__isnull = False) & ~Q(equipment__equipment__vendor__name = ''),
                    then = F('equipment__equipment__vendor__name')
                ),
                default = F('equipment__equipment__brand__name'), output_field = CharField()
            ),
            equipment_name = Concat('equipment__equipment__brand__name', Value(' '), 'equipment__equipment__model__name', Value(' (S/n: '), 'equipment__sn', Value(')')),

        ).filter(id = application_id)[0]
        serializer = ApplicationDetailsSerializer(application)

        history = ApplicationHistoryAPIView().get(request = request._request, application_id = application_id).data

        spares = SparesListAPIView.as_view()(request._request).data

        return Response({'application': serializer.data, 'history': history, 'spares': spares, 'permissions': permissions}, status = status.HTTP_200_OK)

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Редактирование заявки',
        description = '<ol><li>"engineer" - Идентификатор инженера</li><li>"priority" - Идентификатор приоритета заявки</li>\
        <li>"status" - Идентификатор статуса заявки</li><li>"equipment" - Идентификатор оборудования</li></ol>',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = EditApplicationSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EditApplicationSerializer())}
    )
    def put(self, request, application_id, *args, **kwargs):
        if not is_admin_or_engineer(request.user):
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        application = ApplicationModel.objects.get(id = application_id)
        data = request.data.copy()

        if data['equipment'] != application.equipment_id:
            if application.changed:
                data['equipment'] = application.equipment_id
            else:
                equipments = GetClientsEquipmentsAPIView().get(request, client_id = application.client_id).data
                equipment = next((equip for equip in equipments['equipments'] if equip['id'] == data['equipment']), None)
                if not equipment:
                    return Response({'message': 'Оборудование указано некорректно'}, status = status.HTTP_406_NOT_ACCEPTABLE)

        if application.status_id == 1 and 'engineer' in data and data['engineer']:
            data['status'] = 6

        history = []
        params = {
            'id': application_id,
            'url': request.build_absolute_uri(application.get_absolute_url()),
            'status': '',
            'type': 'edit'
        }
        if application.engineer or data['engineer']:
            engineer_id = application.engineer_id if application.engineer_id else data['engineer']

            User = get_user_model()
            engineer = User.objects.get(id = engineer_id)
            engineer_email = engineer.email

        if application.status_id != data['status']:
            app_status = StatusModel.objects.get(id = data['status'])
            history.append({'type': 1, 'text': 'Статус заявки изменен на "' + app_status.name + '"', 'application': application, 'author': request.user})

            AppStatusModel.objects.create(application_id = application_id, status_id = data['status'])

            contact_email = application.contact.email
            params['status'] = app_status
            if send_email(params = params, title = 'Изменение статуса заявки', send_to = [contact_email]):
                text = 'Отправлено сообщение об изменении статуса заявки на "' + app_status.name + '" на адрес электронной почты ' + contact_email
            else:
                text = 'Не удалось отправить сообщение об изменении статуса заявки на "' + app_status.name + '" на адрес электронной почты ' + contact_email
            history.append({'type': 3, 'text': text, 'application': application, 'author': request.user})

            if send_telegram(params = params):
                text = 'Отправлено сообщение об изменении статуса заявки на "' + app_status.name + '" в телеграм-канал'
            else:
                text = 'Не удалось отправить сообщение об изменении статуса заявки на "' + app_status.name + '" в телеграм-канал'
            history.append({'type': 4, 'text': text, 'application': application, 'author': request.user})

            if application.engineer or data['engineer']:
                params['type'] = 'status'

                if send_email(params = params, title = 'Изменение статуса заявки', send_to = [engineer_email]):
                    text = 'Отправлено сообщение об изменении статуса заявки "' + app_status.name + '" на адрес электронной почты ' + engineer_email
                else:
                    text = 'Не удалось отправить сообщение об изменении статуса заявки "' + app_status.name + '" на адрес электронной почты ' + engineer_email

                history.append({'type': 4, 'text': text, 'application': application, 'author': request.user})

        if application.priority_id != data['priority']:
            app_priority = AppPriorityModel.objects.get(id = data['priority'])
            history.append({'type': 2, 'text': 'Изменен приоритет заявки на "' + app_priority.name + '"', 'application': application, 'author': request.user})

            if application.engineer or data['engineer']:
                params['type'] = 'priority'
                params['status'] = app_priority
                if send_email(params = params, title = 'Изменение приоритета заявки', send_to = [engineer_email]):
                    text = 'Отправлено сообщение об изменении приоритета заявки "' + app_priority.name + '" на адрес электронной почты ' + engineer_email
                else:
                    text = 'Не удалось отправить сообщение об изменении приоритета заявки "' + app_priority.name + '" на адрес электронной почты ' + engineer_email
                history.append({'type': 3, 'text': text, 'application': application, 'author': request.user})

                if send_telegram(params = params):
                    text = 'Отправлено сообщение об изменении приоритета заявки на "' + app_status.name + '" в телеграм-канал'
                else:
                    text = 'Не удалось отправить сообщение об изменении приоритета заявки на "' + app_status.name + '" в телеграм-канал'
                history.append({'type': 4, 'text': text, 'application': application, 'author': request.user})

        if 'engineer' in data and not application.engineer:
            history.append({'type': 2, 'text': 'Назначен инженер "' + str(engineer) + '"', 'application': application, 'author': request.user})

            params['type'] = 'engineer'
            params['status'] = engineer
            if send_email(params = params, title = 'Назначен инженер', send_to = [engineer_email]):
                text = 'Отправлено сообщение о назначении инженера "' + str(engineer) + '" на адрес электронной почты ' + engineer_email
            else:
                text = 'Не удалось отправить сообщение о назначении инженера "' + str(engineer) + '" на адрес электронной почты ' + engineer_email
            history.append({'type': 4, 'text': text, 'application': application, 'author': request.user})

            if send_telegram(params = params):
                text = 'Отправлено сообщение о назначении инженера "' + str(engineer) + '" в телеграм-канал'
            else:
                text = 'Не удалось отправить сообщение о назначении инженера "' + str(engineer) + '" в телеграм-канал'
            history.append({'type': 4, 'text': text, 'application': application, 'author': request.user})

        history_instance = [AppHistoryModel(**row) for row in history]
        AppHistoryModel.objects.bulk_create(history_instance)

        serializer = EditApplicationSerializer(application, data = data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Удаление заявки',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Объект удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Объект удален'})])}
    )
    def delete(self, request, application_id, *args, **kwargs):
        if not request.user.groups.filter(name = 'Администратор').exists():
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        application = ApplicationModel.objects.annotate(email = F('contact__email')).get(id = application_id)
        email = application.email
        application.delete()

        history = []
        params = {
            'id': application_id,
            'type': 'delete'
        }

        history.append({'type': 1, 'text': 'Заявка удалена', 'number': application_id, 'author': request.user})

        if send_email(params = params, title = 'Изменение статуса заявки', send_to = [email]):
            text = 'Отправлено сообщение об удалении заявки на адрес электронной почты ' + email
        else:
            text = 'Не удалось отправить сообщение об удалении заявки на адрес электронной почты ' + email
        history.append({'type': 3, 'text': text, 'number': application_id, 'author': request.user})

        if send_telegram(params = params):
            text = 'Отправлено сообщение об удалении заявки в телеграм-канал'
        else:
            text = 'Не удалось отправить сообщение об удалении заявки в телеграм-канал'
        history.append({'type': 4, 'text': text, 'number': application_id, 'author': request.user})

        history_instance = [AppHistoryModel(**row) for row in history]
        AppHistoryModel.objects.bulk_create(history_instance)

        return Response({'message': 'Заявка удалена'}, status = status.HTTP_200_OK)

class EditDoneApplicationAPIView(APIView): #Редактирование заявки
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Редактирование заявки. Возвращение в работу закрытой заявки / Согласование закрытия',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
            OpenApiParameter(name = 'status_id', description = 'Идентификатор статуса возврата / согласования заявки', type = int, location = OpenApiParameter.PATH,
                examples = [
                    OpenApiExample('Возвращение в работу', value = 6),
                    OpenApiExample('Согласование закрытия', value = 4),
                ]
            ),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Статус заявки изменен'}, examples = [OpenApiExample('Пример', value = {'message': 'Статус заявки изменен'})])}
    )
    def patch(self, request, application_id, status_id, *args, **kwargs):
        application = ApplicationModel.objects.annotate(
            email = F('contact__email'),
            engineer_email = F('engineer__email'),
            organization_id = F('client__organization_id')
        ).get(id = application_id)

        if not request.user.groups.filter(name = 'Администратор').exists() and not application.organization_id == request.user.organization_id:
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        history = []
        params = {
            'id': application_id,
            'url': request.build_absolute_uri(application.get_absolute_url()),
            'type': 'to_work'
        }

        if not application.status_id in [3, 4]:
            return Response({'message': 'Неверный статус заявки'}, status = status.HTTP_400_BAD_REQUEST)
        if status_id == 6:
            title = 'Заявка возвращена в работу'
            text_part = 'о возвращении заявки в работу'
        elif status_id == 4:
            title = 'Статус заявки № ' + str(application_id) + ' изменен на "Закрыта".'
            params['type'] = 'status'
            text_part = 'об изменении статуса заявки'
        else:
            return Response({'message': 'Неверный статус заявки'}, status = status.HTTP_400_BAD_REQUEST)

        application.status_id = status_id
        application.save()

        AppStatusModel.objects.create(application_id = application_id, status_id = status_id)

        if status_id == 4:
            params['status'] = application.status

        email_list = [application.email, ]
        if application.engineer_email:
            email_list.append(application.engineer_email)

        history.append({'type': 1, 'text': title, 'number': application_id, 'author': request.user})

        if send_email(params = params, title = 'Изменение статуса заявки', send_to = email_list):
            text = 'Отправлено сообщение ' + text_part + ' на адреса электронной почты'
        else:
            text = 'Не удалось отправить сообщение ' + text_part + ' на адреса электронной почты'
        history.append({'type': 3, 'text': text, 'number': application_id, 'author': request.user})

        if status_id == 4:
            params['type'] = 'edit'

        if send_telegram(params = params):
            text = 'Отправлено сообщение ' + text_part + ' в телеграм-канал'
        else:
            text = 'Не удалось отправить сообщение ' + text_part + ' в телеграм-канал'
        history.append({'type': 4, 'text': text, 'number': application_id, 'author': request.user})

        history_instance = [AppHistoryModel(**row) for row in history]
        AppHistoryModel.objects.bulk_create(history_instance)

        return Response({'message': 'Статус заявки изменен'}, status = status.HTTP_200_OK)

class ApplicationDocumentsAPIView(APIView): #Редактирование заявки. Добавление файлов
    permission_classes = [IsAuthenticated,]
    parser_classes = [MultiPartParser]# FileUploadParser

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Редактирование заявки. Добавление файлов',
        description = '"documents" - Список файлов, где<ol><li>"document" - файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b></li><li>"name" - Наименование файла (если не указано, то значение будет равно имени файла)</li></ol>',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = {'application/json': AddAppDocumentsSerializer()},
        responses = {(200, 'application/json'): OpenApiResponse(response = AddAppDocumentsSerializer())}
    )
    def post(self, request, application_id, *args, **kwargs):
        if not is_admin_or_engineer(request.user):
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        application = ApplicationModel.objects.get(id = application_id)
        serializer = AddAppDocumentsSerializer(application, data = request.data, context = {'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

class ApplicationCommentsAPIView(APIView): #Редактирование заявки. Добавление комментариев
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Редактирование заявки. Список комментариев',
        description = '"comments" - Список комментариев заявки, где<ol><li>"id" - Идентификатор комментария</li>\
        <li>"text" - Текст комментария</li><li>"formatted_date" - Дата создания</li><li>"author_name" - ФИО автора</li></ol>',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = CommentsListSerializer())}
    )
    def get(self, request, application_id, *args, **kwargs):
        application = ApplicationModel.objects.annotate(organization_id = F('client__organization_id')).get(id = application_id)

        if not is_admin_or_engineer(request.user) and not application.organization_id == request.user.organization_id:
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        params = {'application_id': application_id}
        if not is_admin_or_engineer(request.user):
            params['hide'] = False

        comments = AppCommentModel.objects.filter(**params).annotate(
            formatted_date = Func(
                Func(F('pubdate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                Value('%d.%m.%Y %H:%i'), function = 'DATE_FORMAT', output_field = CharField()
            ),
            author_name = Trim(Case(
                When(
                    Q(author__last_name__isnull = False),
                    then = Concat('author__first_name', Value(' '), 'author__last_name')
                ),
                When(author__last_name = '', then = F('author__email')),
                default = Value(''), output_field = CharField()
            ))
        ).values('id', 'text', 'formatted_date', 'author_name')

        return Response({'comments': comments}, status = status.HTTP_200_OK)

    @extend_schema(
        tags = ['Заявки (Done)'],
        summary = 'Редактирование заявки. Добавление комментария',
        description = '<ol><li>"text" - Текст комментария</li><li>"hide" - true, если комментарий не должен отображаться у заказчика</li></ol>',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = AppCommentSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = AppCommentSerializer())}
    )
    def post(self, request, application_id, *args, **kwargs):
        application = ApplicationModel.objects.annotate(organization_id = F('client__organization_id')).get(id = application_id)

        if not is_admin_or_engineer(request.user) and not application.organization_id == request.user.organization_id:
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        serializer = AppCommentSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save(application = application, author = request.user)
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

@extend_schema(
    tags = ['Заявки (Done)']
)
class AppCommentAPIView(APIView):
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        summary = 'Редактирование заявки. Редактирование комментария',
        description = '<ol><li>"text" - Текст комментария</li><li>"hide" - true, если комментарий не должен отображаться у заказчика</li></ol>',
        parameters = [
            OpenApiParameter(name = 'comment_id', description = 'Идентификатор комментария', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = AppCommentSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = AppCommentSerializer())}
    )
    def put(self, request, comment_id, *args, **kwargs):
        comment = GetComment(comment_id)
        if not is_admin_or_engineer(request.user) and not request.user == comment.author:
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        serializer = AppCommentSerializer(comment, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary = 'Редактирование заявки. Удаление комментария',
        parameters = [
            OpenApiParameter(name = 'comment_id', description = 'Идентификатор комментария', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Объект удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Элемент удален'})])}
    )
    def delete(self, request, comment_id, *args, **kwargs):
        comment = GetComment(comment_id)
        if not is_admin_or_engineer(request.user) and not request.user == comment.author:
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        comment.delete()
        return Response({'message': 'Элемент удален'}, status = status.HTTP_200_OK)

def GetComment(comment_id):
    try:
        return AppCommentModel.objects.get(id = comment_id)
    except:
        raise APIException('Комментарий с id = ' + str(comment_id) + ' не найден')

@extend_schema(
    tags = ['Заявки (Done)']
)
class ApplicationHistoryAPIView(APIView):
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        summary = 'История действий: Изменение статуса, Изменение значения поля, Списание запчасти, Возврат запчасти',
        description = '<ol><li>"id" - Идентификатор комментария / события</li><li>"pubdate" - Дата создания</li>\
        <li>"text" - Текст</li><li>"author_name" - ФИО автора, если указано, иначе e-mail</li><li>"formatted_date" - Форматированная дата создания</li>\
        <li>"record_type" - Тип записи ("comment" - комментарий, "status" - событие)</li><li>"hide" - True, если запись - комментарий, и его нужно скрыть от пользователя</li></ol>',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = HistorySerializer())}
    )
    def get(self, request, application_id, *args, **kwargs):
        if not is_admin_or_engineer(request.user):
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        statuses = AppHistoryModel.objects.filter(Q(application_id = application_id) & Q(Q(type = 1) | Q(type = 2) | Q(type = 5) | Q(type = 6)))\
            .annotate(
                author_name = Trim(
                    Case(
                        When(
                            Q(author__last_name__isnull = False) & ~Q(author__last_name = ''),
                            then = Concat('author__last_name', Value(' '), 'author__first_name', output_field = CharField())
                        ),
                        default = F('author__email'), output_field = CharField()
                    )
                ),
                formatted_date = Func(
                    Func(F('pubdate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                    Value('%d.%m.%Y %H:%i'), function = 'DATE_FORMAT', output_field = CharField()
                ),
                record_type = Value('status'),
                hide = Value('')
            )\
            .order_by('pubdate').values('id', 'pubdate', 'formatted_date', 'text', 'author_name', 'record_type', 'hide')

        comments = AppCommentModel.objects.filter(application_id = application_id)\
            .annotate(
                author_name = Trim(Concat('author__last_name', Value(' '), 'author__first_name', output_field = CharField())),
                formatted_date = Func(
                    Func(F('pubdate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                    Value('%d.%m.%Y %H:%i'), function = 'DATE_FORMAT', output_field = CharField()
                ),
                record_type = Value('comment')
            )\
            .order_by('pubdate').values('id', 'pubdate', 'formatted_date', 'text', 'author_name', 'record_type', 'hide')
        history = sorted(chain(statuses, comments), key=lambda item: item['pubdate'], reverse = True)

        return Response(history, status = status.HTTP_200_OK)

@extend_schema(
    tags = ['Заявки (Done)']
)
class ApplicationLogsAPIView(APIView): #Логи заявки
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        summary = 'Логи заявки',
        description = '<ol><li>"id" - Идентификатор комментария / события</li><li>"pubdate" - Дата создания</li>\
        <li>"text" - Текст</li><li>"author_name" - ФИО автора, если указано, иначе e-mail</li><li>"formatted_date" - Форматированная дата создания</li>\
        <li>"record_type" - Тип записи ("comment" - комментарий, "status" - событие)</li><li>"hide" - True, если запись - комментарий, и его нужно скрыть от пользователя</li></ol>',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = HistorySerializer())}
    )
    def get(self, request, application_id, *args, **kwargs):
        if not is_admin_or_engineer(request.user):
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        statuses = AppHistoryModel.objects.filter(application_id = application_id)\
            .annotate(
                author_name = Trim(
                    Case(
                        When(
                            Q(author__last_name__isnull = False) & ~Q(author__last_name = ''),
                            then = Concat('author__last_name', Value(' '), 'author__first_name', output_field = CharField())
                        ),
                        default = F('author__email'), output_field = CharField()
                    )
                ),
                formatted_date = Func(
                    Func(F('pubdate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                    Value('%d.%m.%Y %H:%i'), function = 'DATE_FORMAT', output_field = CharField()
                ),
                record_type = Value('status'),
                hide = Value('')
            )\
            .order_by('pubdate').values('id', 'pubdate', 'formatted_date', 'text', 'author_name', 'record_type', 'hide')

        comments = AppCommentModel.objects.filter(application_id = application_id)\
            .annotate(
                author_name = Trim(Concat('author__last_name', Value(' '), 'author__first_name', output_field = CharField())),
                formatted_date = Func(
                    Func(F('pubdate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                    Value('%d.%m.%Y %H:%i'), function = 'DATE_FORMAT', output_field = CharField()
                ),
                record_type = Value('comment')
            )\
            .order_by('pubdate').values('id', 'pubdate', 'formatted_date', 'text', 'author_name', 'record_type', 'hide')
        history = sorted(chain(statuses, comments), key=lambda item: item['pubdate'], reverse = True)

        return Response(history, status = status.HTTP_200_OK)

@extend_schema(
    tags = ['Заявки (Done)']
)
class AppSpareAPIView(APIView): #Запчасти заявки
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        summary = 'Редактирование заявки. Списание запчастей',
        description = '"spare" - Идентификатор запчасти',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = AppSpareSerializer(many = True),
        responses = {(200, 'application/json'): OpenApiResponse(response = AppSpareSerializer(many = True))}
    )
    def post(self, request, application_id, *args, **kwargs):
        application = ApplicationModel.objects.get(id = application_id)

        if not is_admin_or_engineer(request.user) and not application.engineer == request.user:
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        serializer = AppSpareSerializer(data = request.data, many = True)
        if serializer.is_valid():
            serializer.save(application = application, author = request.user)
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary = 'Редактирование заявки. Отмена списания запчастей',
        description = '"appeqspare" - Идентификатор записи запчасти в заявке',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = EditAppSpareSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EditAppSpareSerializer())}
    )
    def put(self, request, application_id, *args, **kwargs):
        application = ApplicationModel.objects.get(id = application_id)
        if not is_admin_or_engineer(request.user) and not application.engineer == request.user:
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        serializer = EditAppSpareSerializer(data = request.data)
        if serializer.is_valid():
            if serializer.data['appeqspare']:
                spares = AppSpareModel.objects.filter(id__in = serializer.data['appeqspare'])
                for spare in spares: #Чтобы отрабатывал метод delete модели
                    spare.delete()

            return Response({'message': 'Списание запчастей отменено'}, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

@extend_schema(
    tags = ['Заявки (Done)']
)
class ApplicationDetailsAPIView(APIView): #Детальная информация о заявке
    permission_classes = [IsAuthenticated,]

    @extend_schema(
        summary = 'Детальная информация о заявке',
        description = '<p>"application" - Данные заявки, где</p><ol><li>"id" - Идентификатор / номер заявки</li><li>"formatted_date" - Дата создания</li>\
        <li>"organization_id" - Идентификатор организации</li>\
        <li>"status_id" - Идентификатор статуса заявки <b>(Доступен только персоналу)</b></li>\
        <li>"status_name" - Наименование статуса заявки <b>(Доступен только персоналу)</b></li>\
        <li>"priority_id" - Идентификатор приоритета заявки <b>(Доступен только персоналу)</b></li>\
        <li>"priority_name" - Наименование приоритета заявки <b>(Доступен только персоналу)</b></li>\
        <li>"engineer_id" - Идентификатор инженера</li><li>"engineer_name" - ФИО инженера, если указано, иначе e-mail</li>\
        <li>"contact_id" - Идентификатор контактного лица</li><li>"contact_name" - ФИО контактного лица</li><li>"contact_email" - e-mail контактного лица</li>\
        <li>"contact_phone" - Телефон контактного лица</li><li>"equipment_id" - Идентификатор оборудования</li><li>"equipment_name" - Наименование оборудования</li>\
        <li>"support_id" - Идентификатор уровня поддержки</li><li>"support_level" - Наименование уровня поддержки</li>\
        <li>"vendor_name" - Наименование вендора <b>(Доступен только персоналу)</b></li><li>"problem" - Описание проблемы</li>\
        <li>"documents" - Список прикрепленных файлов, где<ul><li>"document" - файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b></li><li>"name" - Наименование файла (если не указано, то значение будет равно имени файла)</li></ul></li></ol>\
        <p>"history" - Действия заявки, где</p><ol><li>"id" - Идентификатор комментария / события</li>\
        <li>"text" - Текст</li><li>"formatted_date" - Форматированная дата создания</li><li>"author_name" - ФИО автора, если указано, иначе e-mail</li></ol>\
        <p>"permissions" - список условных разрешений пользователя, где</p>\
        <ol><li>"is_admin" - true, если пользователь - администратор</li><li>"is_engineer" - true, если пользователь - инженер</li><li>"is_staff" - true, если пользователь - персонал (администратор или инженер)</li></ol>',
        parameters = [
            OpenApiParameter(name = 'application_id', description = 'Идентификатор заявки', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = ApplicationInfoSerializer())}
    )
    def get(self, request, application_id, *args, **kwargs):
        permissions = {
            'is_admin': request.user.groups.filter(name = 'Администратор').exists(),
            'is_engineer': request.user.groups.filter(name = 'Инженер').exists(),
            'is_staff': is_admin_or_engineer(request.user)
        }

        application = ApplicationModel.objects.annotate(
            organization_id = F('client__organization_id'),
            formatted_date = Func(
                Func(F('pubdate'), Value('+00:00'), Value('+03:00'), function = 'CONVERT_TZ', output_field = CharField()),
                Value('%d.%m.%Y %H:%i'), function = 'DATE_FORMAT', output_field = CharField()
            ),
            status_name = F('status__name'),
            priority_name = F('priority__name'),
            engineer_name = Case(
                When(
                    Q(engineer__isnull = False) & Q(engineer__last_name__isnull = False) & ~Q(engineer__last_name = ''),
                    then = Concat('engineer__first_name', Value(' '), 'engineer__last_name')
                ),
                When(engineer__last_name = '', then = F('engineer__email')),
                default = Value(''), output_field = CharField()),
            contact_name = F('contact__fio'),
            contact_email = F('contact__email'),
            contact_phone = Func(F('contact__phone'), Value(''), function = 'IFNULL', output_field = CharField()),
            equipment_name = Concat('equipment__equipment__brand__name', Value(' '), 'equipment__equipment__model__name', Value(' (S/n: '), 'equipment__sn', Value(')')),
            support_id = F('equipment__support_id'),
            support_name = F('equipment__support__name'),
            vendor_name = Func(F('equipment__equipment__vendor__name'), F('equipment__equipment__brand__name'), function = 'IFNULL', output_field = CharField())
        ).prefetch_related('documents').get(id = application_id)

        if not application.organization_id == request.user.organization_id and not permissions['is_staff']:
            return Response({'message': 'Недостаточно прав'}, status = status.HTTP_403_FORBIDDEN)

        serializer = ApplicationSerializer(application)
        application_data = serializer.data
        if not permissions['is_staff']:
            del application_data['priority_id']
            del application_data['priority_name']
            del application_data['support_id']
            del application_data['support_name']
            del application_data['vendor_name']

        history = ApplicationCommentsAPIView().get(request = request, application_id = application_id).data

        return Response({'application': application_data, 'history': history['comments'], 'permissions': permissions}, status = status.HTTP_200_OK)

def GetApplication(application_id):
    try:
        return ApplicationModel.objects.get(id = application_id)
    except:
        raise APIException('Заявка с id = ' + str(application_id) + ' не найдена')
