from django.db.models import F
from django.http import HttpResponse

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser

<<<<<<< HEAD
from drf_spectacular.utils import extend_schema
=======
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiExample, OpenApiResponse
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10

from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import NamedStyle, Side, Font, Border, Alignment

from equipments.api.serializers import *

from equipments.models import EquipmentModel, TypeModel, VendorModel, BrandModel, ModelModel

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Оборудование (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentsListAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список оборудования (Done)'],
=======
        summary = 'Список оборудования',
        description = '<ol><li>"id" - Идентификатор оборудования</li><li>"type_name" - Наименование типа оборудования</li>\
        <li>"vendor_name" - Наименование вендора оборудования</li><li>"brand_name" - Наименование бренда оборудования</li>\
        <li>"model_name" - Наименование модели оборудования</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentsListSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):

        list = EquipmentModel.objects.all().annotate(
            type_name = F('type__name'),
            vendor_name = F('vendor__name'),
            brand_name = F('brand__name'),
            model_name = F('model__name')
        )

        serializer = EquipmentsListSerializer(list, many = True)
        return Response(serializer.data, status = status.HTTP_200_OK)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список оборудования. Добавление (Done)'],
        request = EditEquipmentSerializer()
=======
        summary = 'Добавление оборудования',
        description = '<ol><li>"type" - Идентификатор типа оборудования</li>\
        <li>"vendor" - Идентификатор вендора оборудования</li><li>"brand" - Идентификатор бренда оборудования</li>\
        <li>"model" - Идентификатор модели оборудования</li></ol>',
        request = EditEquipmentSerializer(),
        responses = {(201, 'application/json'): OpenApiResponse(response = EditEquipmentSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EditEquipmentSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Оборудование (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EditEquipmentAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список оборудования. Изменение (Done)'],
        request = EditEquipmentSerializer()
=======
        summary = 'Изменение оборудования',
        description = '<ol><li>"type" - Идентификатор типа оборудования</li>\
        <li>"vendor" - Идентификатор вендора оборудования</li><li>"brand" - Идентификатор бренда оборудования</li>\
        <li>"model" - Идентификатор модели оборудования</li></ol>',
        parameters = [
            OpenApiParameter(name = 'equipment_id', description = 'Идентификатор оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = EditEquipmentSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EditEquipmentSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def put(self, request, equipment_id, *args, **kwargs):
        equipment = GetEquipment(equipment_id)

        serializer = EditEquipmentSerializer(equipment, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список оборудования. Удаление (Done)'],
=======
        summary = 'Удаление оборудования',
        parameters = [
            OpenApiParameter(name = 'equipment_id', description = 'Идентификатор оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элемент удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Элемент удален'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def delete(self, request, equipment_id, *args, **kwargs):
        equipment = GetEquipment(equipment_id)
        equipment.delete()
        return Response({'message': 'Элемент удален'}, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Оборудование (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentsDeleteAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список оборудования. Удаление нескольких (Done)'],
        request = EquipmentsDeleteSerializer(),
=======
        summary = 'Удаление списка оборудования',
        description = 'Используется метод POST, т.к. методом DELETE невозможно отправить тело запроса\
        <p>"id" - Список идентификаторов оборудования { "id": [1, 2, 3] }</p>',
        request = EquipmentsDeleteSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элементы удалены'}, examples = [OpenApiExample('Пример', value = {'message': 'Элементы удалены'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EquipmentsDeleteSerializer({'id': request.data.getlist('id')})
        if serializer:
            EquipmentModel.objects.filter(id__in = serializer.data.get('id')).delete()
            return Response({'message': 'Элементы удалены'}, status = status.HTTP_200_OK)
        return Response({'error': 'Некорректные данные'}, status = status.HTTP_400_BAD_REQUEST)

def GetEquipment(equipment_id):
    try:
        return EquipmentModel.objects.get(id = equipment_id)
    except:
        return Response({'error': 'Объект не найден'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
#Типы оборудования
=======
@extend_schema(
    tags = ['Типы оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentTypesListAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список типов оборудования (Done)'],
    )
    def get(self, request, *args, **kwargs):

        list = TypeModel.objects.all()

=======
        summary = 'Список типов оборудования',
        description = '<ol><li>"id" - Идентификатор типа оборудования</li><li>"name" - Наименование типа оборудования</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentTypesSerializer(many = True))}
    )
    def get(self, request, *args, **kwargs):
        list = TypeModel.objects.all()
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
        serializer = EquipmentTypesSerializer(list, many = True)
        return Response(serializer.data, status = status.HTTP_200_OK)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список типов оборудования. Добавление (Done)'],
        request = EquipmentTypesSerializer()
=======
        summary = 'Добавление типа оборудования',
        description = '"name" - Наименование типа оборудования',
        request =  EquipmentTypesSerializer(),
        responses = {(201, 'application/json'): OpenApiResponse(response = EquipmentTypesSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EquipmentTypesSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Типы оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EditEquipmentTypeAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список типов оборудования. Изменение (Done)'],
        request = EquipmentTypesSerializer()
=======
        summary = 'Изменение типа оборудования',
        description = '"name" - Наименование типа оборудования',
        parameters = [
            OpenApiParameter(name = 'type_id', description = 'Идентификатор типа оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = EquipmentTypesSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentTypesSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def put(self, request, type_id, *args, **kwargs):
        type = GetEquipmentType(type_id)

        serializer = EquipmentTypesSerializer(type, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список типов оборудования. Удаление (Done)'],
=======
        summary = 'Удаление типа оборудования',
        parameters = [
            OpenApiParameter(name = 'type_id', description = 'Идентификатор типа оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элемент удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Элемент удален'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def delete(self, request, type_id, *args, **kwargs):
        type = GetEquipmentType(type_id)
        type.delete()
        return Response({'message': 'Элемент удален'}, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Типы оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentTypesDeleteAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список типов оборудования. Удаление нескольких (Done)'],
        request = EquipmentTypesDeleteSerializer(),
=======
        summary = 'Удаление списка типов оборудования',
        description = 'Используется метод POST, т.к. методом DELETE невозможно отправить тело запроса\
        <p>"id" - Список идентификаторов оборудования { "id": [1, 2, 3] }</p>',
        request = EquipmentTypesDeleteSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элементы удалены'}, examples = [OpenApiExample('Пример', value = {'message': 'Элементы удалены'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EquipmentTypesDeleteSerializer({'id': request.data.getlist('id')})
        if serializer:
            TypeModel.objects.filter(id__in = serializer.data.get('id')).delete()
            return Response({'message': 'Элементы удалены'}, status = status.HTTP_200_OK)
        return Response({'error': 'Некорректные данные'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Типы оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentTypesExcelAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список типов оборудования. Экспорт (Done)'],
=======
        summary = 'Экспорт списка типов оборудования в Excel',
        description = 'Response - файл с данными',
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):
        types = EquipmentTypesListAPIView().get(request = request).data

        stream = ExportItems(types, ['Наименование',], 'Типы оборудования')

        response = HttpResponse(content = stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=types.xlsx'
        return response

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список типов оборудования. Импорт (Done)'],
        request = EquipmentImportSerializer()
=======
        summary = 'Импорт списка типов оборудования из Excel',
        description = '"excel" - xlsx-файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b>',
        request = EquipmentImportSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentTypesSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        records = LoadSimpleItems(request.data.get('excel'))

        list = [TypeModel(**vals) for vals in records]
        TypeModel.objects.bulk_create(list, ignore_conflicts = True)
        return Response(records, status = status.HTTP_200_OK)

def GetEquipmentType(type_id):
    try:
        return TypeModel.objects.get(id = type_id)
    except:
        return Response({'error': 'Объект не найден'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
#Вендоры оборудования
=======
@extend_schema(
    tags = ['Вендоры оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentVendorsListAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список вендоров оборудования (Done)'],
    )
    def get(self, request, *args, **kwargs):

        list = VendorModel.objects.all()

=======
        summary = 'Список вендоров оборудования',
        description = '<ol><li>"id" - Идентификатор вендора оборудования</li><li>"name" - Наименование вендора оборудования</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentVendorsSerializer(many = True))}
    )
    def get(self, request, *args, **kwargs):
        list = VendorModel.objects.all()
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
        serializer = EquipmentVendorsSerializer(list, many = True)
        return Response(serializer.data, status = status.HTTP_200_OK)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список вендоров оборудования. Добавление (Done)'],
        request = EquipmentVendorsSerializer()
=======
        summary = 'Добавление вендора оборудования',
        description = '"name" - Наименование вендора оборудования',
        request = EquipmentVendorsSerializer(),
        responses = {(201, 'application/json'): OpenApiResponse(response = EquipmentVendorsSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EquipmentVendorsSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Вендоры оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EditEquipmentVendorAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список вендоров оборудования. Изменение (Done)'],
        request = EquipmentVendorsSerializer()
=======
        summary = 'Изменение вендора оборудования',
        description = '"name" - Наименование вендора оборудования',
        parameters = [
            OpenApiParameter(name = 'vendor_id', description = 'Идентификатор вендора оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = EquipmentVendorsSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentVendorsSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def put(self, request, vendor_id, *args, **kwargs):
        vendor = GetEquipmentVendor(vendor_id)

        serializer = EquipmentVendorsSerializer(vendor, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список вендоров оборудования. Удаление (Done)'],
=======
        summary = 'Удаление вендора оборудования',
        parameters = [
            OpenApiParameter(name = 'vendor_id', description = 'Идентификатор вендора оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элемент удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Элемент удален'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def delete(self, request, vendor_id, *args, **kwargs):
        vendor = GetEquipmentVendor(vendor_id)
        vendor.delete()
        return Response({'message': 'Элемент удален'}, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Вендоры оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentVendorsDeleteAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список вендоров оборудования. Удаление нескольких (Done)'],
        request = EquipmentVendorsDeleteSerializer(),
=======
        summary = 'Удаление списка вендоров оборудования',
        description = 'Используется метод POST, т.к. методом DELETE невозможно отправить тело запроса\
        <p>"id" - Список идентификаторов вендоров { "id": [1, 2, 3] }</p>',
        request = EquipmentVendorsDeleteSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элементы удалены'}, examples = [OpenApiExample('Пример', value = {'message': 'Элементы удалены'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EquipmentVendorsDeleteSerializer({'id': request.data.getlist('id')})
        if serializer:
            VendorModel.objects.filter(id__in = serializer.data.get('id')).delete()
            return Response({'message': 'Элементы удалены'}, status = status.HTTP_200_OK)
        return Response({'error': 'Некорректные данные'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Вендоры оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentVendorsExcelAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список вендоров оборудования. Экспорт (Done)'],
=======
        summary = 'Экспорт списка вендоров оборудования в Excel',
        description = 'Response - файл с данными',
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):
        vendors = EquipmentVendorsListAPIView().get(request = request).data

        stream = ExportItems(vendors, ['Наименование',], 'Список вендоров оборудования')

        response = HttpResponse(content = stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=vendors.xlsx'
        return response

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список вендоров оборудования. Импорт (Done)'],
        request = EquipmentImportSerializer()
=======
        summary = 'Импорт списка вендоров оборудования из Excel',
        description = '"excel" - xlsx-файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b>',
        request = EquipmentImportSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentVendorsSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        records = LoadSimpleItems(request.data.get('excel'))

        list = [VendorModel(**vals) for vals in records]
        VendorModel.objects.bulk_create(list, ignore_conflicts = True)
        return Response(records, status = status.HTTP_200_OK)

def GetEquipmentVendor(vendor_id):
    try:
        return VendorModel.objects.get(id = vendor_id)
    except:
        return Response({'error': 'Объект не найден'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
#Бренды оборудования
=======
@extend_schema(
    tags = ['Бренды оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentBrandsListAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список брендов оборудования (Done)'],
=======
        summary = 'Список брендов оборудования',
        description = '<ol><li>"id" - Идентификатор бренда оборудования</li><li>"name" - Наименование бренда оборудования</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentBrandsSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):

        list = BrandModel.objects.all()

        serializer = EquipmentBrandsSerializer(list, many = True)
        return Response(serializer.data, status = status.HTTP_200_OK)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список брендов оборудования. Добавление (Done)'],
        request = EquipmentBrandsSerializer()
=======
        summary = 'Добавление бренда оборудования',
        description = '"name" - Наименование бренда оборудования',
        request = EquipmentBrandsSerializer(),
        responses = {(201, 'application/json'): OpenApiResponse(response = EquipmentBrandsSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EquipmentBrandsSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Бренды оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EditEquipmentBrandAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список брендов оборудования. Изменение (Done)'],
        request = EquipmentBrandsSerializer()
=======
        summary = 'Изменение бренда оборудования',
        description = '"name" - Наименование бренда оборудования',
        parameters = [
            OpenApiParameter(name = 'brand_id', description = 'Идентификатор бренда оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = EquipmentBrandsSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentBrandsSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def put(self, request, brand_id, *args, **kwargs):
        brand = GetEquipmentBrand(brand_id)

        serializer = EquipmentBrandsSerializer(brand, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список брендов оборудования. Удаление (Done)'],
=======
        summary = 'Удаление бренда оборудования',
        parameters = [
            OpenApiParameter(name = 'brand_id', description = 'Идентификатор бренда оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элемент удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Элемент удален'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def delete(self, request, brand_id, *args, **kwargs):
        brand = GetEquipmentBrand(brand_id)
        brand.delete()
        return Response({'message': 'Элемент удален'}, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Бренды оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentBrandsDeleteAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список брендов оборудования. Удаление нескольких (Done)'],
        request = EquipmentBrandsDeleteSerializer(),
=======
        summary = 'Удаление списка брендов оборудования',
        description = 'Используется метод POST, т.к. методом DELETE невозможно отправить тело запроса\
        <p>"id" - Список идентификаторов брендов { "id": [1, 2, 3] }</p>',
        request = EquipmentBrandsDeleteSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элементы удалены'}, examples = [OpenApiExample('Пример', value = {'message': 'Элементы удалены'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EquipmentBrandsDeleteSerializer({'id': request.data.getlist('id')})
        if serializer:
            BrandModel.objects.filter(id__in = serializer.data.get('id')).delete()
            return Response({'message': 'Элементы удалены'}, status = status.HTTP_200_OK)
        return Response({'error': 'Некорректные данные'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Бренды оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentBrandsExcelAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список брендов оборудования. Экспорт (Done)'],
=======
        summary = 'Экспорт списка вендоров оборудования в Excel',
        description = 'Response - файл с данными',
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):
        brands = EquipmentBrandsListAPIView().get(request = request).data

        stream = ExportItems(brands, ['Наименование',], 'Список брендов оборудования')

        response = HttpResponse(content = stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=brands.xlsx'
        return response

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список брендов оборудования. Импорт (Done)'],
        request = EquipmentImportSerializer()
=======
        summary = 'Импорт списка брендов оборудования из Excel',
        description = '"excel" - xlsx-файл <b>Внимание! Тип параметра - File {"lastModified": 0, "name": "string", "size": 0, "type": "string", ...}</b>',
        request = EquipmentImportSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentBrandsSerializer(many = True))}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        records = LoadSimpleItems(request.data.get('excel'))

        list = [BrandModel(**vals) for vals in records]
        BrandModel.objects.bulk_create(list, ignore_conflicts = True)
        return Response(records, status = status.HTTP_200_OK)

<<<<<<< HEAD

=======
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
def GetEquipmentBrand(brand_id):
    try:
        return BrandModel.objects.get(id = brand_id)
    except:
        return Response({'error': 'Объект не найден'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
#Модели оборудования
=======
@extend_schema(
    tags = ['Модели оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentModelsListAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список моделей оборудования (Done)'],
    )
    def get(self, request, *args, **kwargs):

=======
        summary = 'Список моделей оборудования',
        description = '<ol><li>"id" - Идентификатор модели оборудования</li><li>"name" - Наименование модели оборудования</li>\
        <li>"brand_name" - Наименование бренда</li><li>"vendor_name" - Наименование вендора</li></ol>',
        responses = {(200, 'application/json'): OpenApiResponse(response = EquipmentModelsSerializer(many = True))}
    )
    def get(self, request, *args, **kwargs):
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
        list = ModelModel.objects.all().annotate(vendor_name = F('vendor__name'), brand_name = F('brand__name'))
        serializer = EquipmentModelsSerializer(list, many = True)
        return Response(serializer.data, status = status.HTTP_200_OK)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список моделей оборудования. Добавление (Done)'],
        request = EditEquipmentModelSerializer()
=======
        summary = 'Добавление модели оборудования',
        description = '<ol><li>"name" - Наименование модели оборудования</li>\
        <li>"brand" - Идентификатор бренда</li><li>"vendor" - Идентификатор вендора</li></ol>',
        request = EditEquipmentModelSerializer(),
        responses = {(201, 'application/json'): OpenApiResponse(response = EditEquipmentModelSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EditEquipmentModelSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Модели оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EditEquipmentModelAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список моделей оборудования. Изменение (Done)'],
        request = EditEquipmentModelSerializer()
=======
        summary = 'Изменение модели оборудования',
        description = '<ol><li>"name" - Наименование модели оборудования</li>\
        <li>"brand" - Идентификатор бренда</li><li>"vendor" - Идентификатор вендора</li></ol>',
        parameters = [
            OpenApiParameter(name = 'model_id', description = 'Идентификатор модели оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        request = EditEquipmentModelSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = EditEquipmentModelSerializer())}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def put(self, request, model_id, *args, **kwargs):
        model = GetEquipmentModel(model_id)

        serializer = EditEquipmentModelSerializer(model, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_200_OK)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список моделей оборудования. Удаление (Done)'],
=======
        summary = 'Удаление модели оборудования',
        parameters = [
            OpenApiParameter(name = 'model_id', description = 'Идентификатор модели оборудования', type = int, required = True, location = OpenApiParameter.PATH),
        ],
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элемент удален'}, examples = [OpenApiExample('Пример', value = {'message': 'Элемент удален'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def delete(self, request, model_id, *args, **kwargs):
        model = GetEquipmentModel(model_id)
        model.delete()
        return Response({'message': 'Элемент удален'}, status = status.HTTP_200_OK)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Модели оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentModelsDeleteAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список моделей оборудования. Удаление нескольких (Done)'],
        request = EquipmentModelsDeleteSerializer(),
=======
        summary = 'Удаление списка моделей оборудования',
        description = 'Используется метод POST, т.к. методом DELETE невозможно отправить тело запроса\
        <p>"id" - Список идентификаторов моделей { "id": [1, 2, 3] }</p>',
        request = EquipmentModelsDeleteSerializer(),
        responses = {(200, 'application/json'): OpenApiResponse(response = {'message': 'Элементы удалены'}, examples = [OpenApiExample('Пример', value = {'message': 'Элементы удалены'})])}
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def post(self, request, *args, **kwargs):
        serializer = EquipmentModelsDeleteSerializer({'id': request.data.getlist('id')})
        if serializer:
            ModelModel.objects.filter(id__in = serializer.data.get('id')).delete()
            return Response({'message': 'Элементы удалены'}, status = status.HTTP_200_OK)
        return Response({'error': 'Некорректные данные'}, status = status.HTTP_400_BAD_REQUEST)

<<<<<<< HEAD
=======
@extend_schema(
    tags = ['Модели оборудования (Done)'],
)
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
class EquipmentModelsExcelAPIView(APIView):
    permission_classes = [IsAdminUser,]

    @extend_schema(
<<<<<<< HEAD
        tags = ['Список моделей оборудования. Экспорт (Done)'],
=======
        summary = 'Экспорт списка моделей оборудования в Excel',
        description = 'Response - файл с данными',
>>>>>>> 0b86e9e586987b8a392d3f43c66c2fbb91b80e10
    )
    def get(self, request, *args, **kwargs):
        models = EquipmentModelsListAPIView().get(request = request).data

        stream = ExportItems(models, ['Наименование', 'Бренд', 'Вендор'], 'Список моделей оборудования')

        response = HttpResponse(content = stream, content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=models.xlsx'
        return response

def GetEquipmentModel(model_id):
    try:
        return ModelModel.objects.get(id = model_id)
    except:
        return Response({'error': 'Объект не найден'}, status = status.HTTP_400_BAD_REQUEST)

#Общие функции
def ExportItems(items, fields, title):
    wb = Workbook()
    with NamedTemporaryFile(delete = True) as tmp:
        ws = wb.active
        ws.title = title

        bd = Side(style = 'thin', color = '000000')

        general = NamedStyle(name = 'general')
        general.font = Font(name = 'Times New Roman', size = 14)
        general.border = Border(left = bd, top = bd, right = bd, bottom = bd)
        general.alignment = Alignment(vertical = 'center', wrap_text = True)
        wb.add_named_style(general)

        header = NamedStyle(name = 'header')
        header.font = Font(name = 'Times New Roman', size = 14, bold = True)
        header.border = Border(left = bd, top = bd, right = bd, bottom = bd)
        header.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text = True)
        wb.add_named_style(header)

        ws.column_dimensions['A'].width = 10

        ws['A2'] = '#'
        ws['A2'].style = 'header'
        col = 2

        for field in fields:
            cell = ws.cell(row = 2, column = col, value = field)
            ws.column_dimensions[get_column_letter(col)].width = 20
            cell.style = 'header'
            col = col + 1

        for count, item in enumerate(items):
            cell = ws.cell(row = count + 3, column = 1, value = count + 1)
            cell.style = 'general'
            col = 2
            for field in item.keys():
                if not field == 'id':
                    cell = ws.cell(row = count + 3, column = col, value = str(item[field] or ''))
                    cell.style = 'general'
                    col = col + 1

        wb.save(tmp.name)
        tmp.seek(0)
        return tmp.read()

def LoadSimpleItems(excel):
    wb = load_workbook(filename = excel.file)

    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    start = 0
    end = ws.max_row
    records = []

    for m in range(3 , end + 1):
        records.append({'name': ws.cell(m, 2).value})

    return records
